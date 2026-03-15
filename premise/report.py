"""
This module export a summary of scenario to an Excel file.
"""

import os
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd
import yaml
from openpyxl.chart import AreaChart, LineChart, Reference
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder

from . import __version__
from .filesystem_constants import DATA_DIR
from .logger import empty_log_files
from .scenario_summary import (
    BATTERY_SECTORS,
    get_sector_catalog,
    load_report_metadata,
    summarize_sector,
)

LOG_REPORTING_FILEPATH = DATA_DIR / "utils" / "logging" / "reporting.yaml"
# directory for log files
DIR_LOGS = Path.cwd() / "export" / "logs"
DIR_LOG_REPORT = Path.cwd() / "export" / "change reports"

# if DIR_LOG_REPORT folder does not exist
# we create it
if not Path(DIR_LOG_REPORT).exists():
    Path(DIR_LOG_REPORT).mkdir(parents=True, exist_ok=True)


def _series_group_to_wide_df(group: dict) -> pd.DataFrame:
    """Convert a structured scenario-summary group into a wide DataFrame."""

    columns = {}
    for series in group.get("series", []):
        points = {
            point["year"]: point["value"]
            for point in series.get("points", [])
            if "year" in point and "value" in point
        }
        if points:
            columns[series["variable"]] = pd.Series(points)

    if not columns:
        return pd.DataFrame()

    df = pd.DataFrame(columns)
    df = df.rename_axis(index=None)
    df = df.dropna(axis=1, how="all")
    df = df.dropna(axis=0, how="all")
    if not df.empty:
        try:
            df.index = pd.to_numeric(df.index)
        except Exception:
            pass
        df = df.sort_index()
    return df


def _write_wide_df(
    worksheet, start_row: int, start_col: int, df: pd.DataFrame
) -> tuple[int, int]:
    """
    Write a wide DataFrame (index as categories, columns as series)
    to the worksheet starting at (start_row, start_col).
    Returns (n_rows_written, n_cols_written).
    """
    if df is None or df.empty:
        return (0, 0)

    # Compose a table with the index in the first column
    table = df.copy()
    table.insert(0, "__cats__", table.index)

    # Use openpyxl helper to get row-wise representation
    rows = dataframe_to_rows(table, index=False, header=True)

    n_rows = 0
    n_cols = 0
    for r_idx, row_vals in enumerate(rows):
        # openpyxl sometimes yields trailing None-only rows; filter them
        if not any(v is not None for v in row_vals):
            continue
        for c_idx, value in enumerate(row_vals, 0):
            worksheet.cell(
                row=start_row + n_rows, column=start_col + c_idx, value=value
            )
            n_cols = max(n_cols, c_idx + 1)
        n_rows += 1

    return (n_rows, n_cols)


def _add_chart_if_data(
    worksheet,
    sector: str,
    title: str,
    start_row: int,
    start_col: int,
    n_rows: int,
    n_cols: int,
    y_axis_label: str | None,
    with_charts: bool = True,
):
    """
    Add an Area/Line chart using the block written by _write_wide_df.
    Guarded to avoid invalid ranges and empty data.
    """
    if not with_charts:
        return

    # Need at least header row + 1 data row, and at least 1 series column
    if n_rows < 2 or n_cols < 2:
        return

    # Cells:
    # - Column layout is: [__cats__ | series1 | series2 | ...]
    # - Header at start_row, data from start_row+1 to start_row+n_rows-1
    # - Categories are in (start_col), series start at (start_col+1)
    min_col_vals = start_col + 1
    max_col_vals = start_col + (n_cols - 1)
    min_row_vals = start_row  # include header for titles_from_data=True
    max_row_vals = start_row + (n_rows - 1)

    # Safety: ensure ranges make sense
    if max_col_vals < min_col_vals or max_row_vals <= min_row_vals:
        return

    values = Reference(
        worksheet,
        min_col=min_col_vals,
        min_row=min_row_vals,
        max_col=max_col_vals,
        max_row=max_row_vals,
    )
    cats = Reference(
        worksheet,
        min_col=start_col,
        min_row=start_row + 1,
        max_row=max_row_vals,
    )

    # Pick chart type
    if any(x in sector for x in ("generation", "mix", "Transport")):
        chart = AreaChart(grouping="stacked")
    elif "eff" in sector:
        chart = LineChart()
    elif "CCS" in sector:
        chart = AreaChart()
    else:
        chart = LineChart()

    # Configure and attach
    chart.add_data(values, titles_from_data=True)
    chart.set_categories(cats)
    chart.title = title
    if y_axis_label:
        chart.y_axis.title = y_axis_label

    chart.x_axis.majorTickMark = "out"
    chart.x_axis.tickLblPos = "nextTo"
    chart.y_axis.majorTickMark = "out"
    chart.y_axis.tickLblPos = "nextTo"
    chart.x_axis.delete = False
    chart.y_axis.delete = False

    chart.height = 8
    chart.width = 16
    # Anchor next to the table block
    chart.anchor = f"{get_column_letter(start_col + 2)}{start_row + 1}"

    worksheet.add_chart(chart)


# --- generate_summary_report (updated) ---------------------------------------


def generate_summary_report(
    scenarios: list, filename: Path, *, with_charts: bool = True
) -> None:
    """
    Generate a summary report of the scenarios.

    with_charts: set to False for huge runs to reduce memory usage.
    """

    metadata = load_report_metadata()
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    for sector_entry in get_sector_catalog(metadata):
        sector = sector_entry["id"]
        sector_summary = summarize_sector(scenarios, sector, metadata=metadata)
        if not sector_summary["scenarios"]:
            continue

        worksheet = workbook.create_sheet(sector)

        col, row = (1, 1)
        worksheet.cell(column=col, row=row, value=sector_summary["explanation"])

        scenario_list = set()
        last_col_used = 1

        for scenario_data in sector_summary["scenarios"]:
            key = (scenario_data["model"], scenario_data["pathway"])
            if key in scenario_list:
                continue

            if scenario_list:
                col = last_col_used + sector_summary.get("offset", 2)

            row = 3

            header = (
                f"{scenario_data['model'].upper()} - "
                f"{scenario_data['pathway'].upper()}"
            )
            worksheet.cell(column=col, row=row, value=header)
            worksheet.cell(column=col, row=row).font = Font(
                bold=True, size=14, underline="single"
            )
            row += 2

            for group in scenario_data["groups"]:
                worksheet.cell(column=col, row=row, value=group["name"])
                row += 3

                df = _series_group_to_wide_df(group)
                if df.empty:
                    continue

                n_rows, n_cols = _write_wide_df(worksheet, row, col, df)
                title = (
                    f"{group['name']} scenario"
                    if sector in BATTERY_SECTORS
                    else f"{group['name']} - {sector} ({sector_summary.get('label', '')})"
                )
                _add_chart_if_data(
                    worksheet=worksheet,
                    sector=sector,
                    title=title,
                    start_row=row,
                    start_col=col,
                    n_rows=n_rows,
                    n_cols=n_cols,
                    y_axis_label=(
                        sector_summary.get("label")
                        if sector in BATTERY_SECTORS
                        else None
                    ),
                    with_charts=with_charts,
                )

                last_col_used = max(last_col_used, col + n_cols - 1)
                row += n_rows + 2

            scenario_list.add(key)

    workbook.save(filename)


# --- generate_change_report (updated, hardened) -------------------------------


def generate_change_report(source, version, source_type, system_model):
    """
    Generate a change report of the scenarios from the log files.
    """

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    log_filepaths = [
        "premise_dac",
        "premise_biomass",
        "premise_electricity",
        "premise_fuel",
        "premise_heat",
        "premise_battery",
        "premise_wind_turbine",
        "premise_transport",
        "premise_steel",
        "premise_metal",
        "premise_cement",
        "premise_emissions",
        "premise_external_scenarios",
        "premise_mapping",
        "premise_validation",
    ]

    # fetch reporting metadata
    with open(LOG_REPORTING_FILEPATH, encoding="utf-8") as f:
        metadata = yaml.load(f, Loader=yaml.FullLoader)

    worksheet = workbook.create_sheet("Change report")
    worksheet.cell(row=1, column=1, value="Library name")
    worksheet.cell(row=1, column=2, value="Library version")
    worksheet.cell(row=1, column=3, value="Report date")
    worksheet.cell(row=1, column=4, value="Source database")
    worksheet.cell(row=1, column=5, value="Source database format")
    worksheet.cell(row=1, column=6, value="Database version")
    worksheet.cell(row=1, column=7, value="Database system model")
    worksheet.cell(row=2, column=1, value="premise")
    worksheet.cell(row=2, column=2, value=".".join(map(str, __version__)))
    worksheet.cell(row=2, column=3, value=datetime.now())
    worksheet.cell(row=2, column=4, value=source)
    worksheet.cell(row=2, column=5, value=source_type)
    worksheet.cell(row=2, column=6, value=version)
    worksheet.cell(row=2, column=7, value=system_model)

    dim_holder = DimensionHolder(worksheet=worksheet)
    for col in range(worksheet.min_column, worksheet.max_column + 1):
        dim_holder[get_column_letter(col)] = ColumnDimension(
            worksheet, min=col, max=col, width=20
        )
    worksheet.column_dimensions = dim_holder

    for name in log_filepaths:
        fp = Path(DIR_LOGS / name).with_suffix(".log")
        if not fp.is_file():
            continue
        if os.stat(fp).st_size == 0:
            continue

        try:
            df = convert_log_to_excel_file(fp)
        except Exception as exc:
            print(f"Warning: failed to read log file {fp}: {exc}")
            continue

        # Create per-sector sheet
        tab_meta = metadata.get(name, {})
        tab_name = tab_meta.get(
            "tab", fetch_tab_name(name) if isinstance(name, str) else name
        )
        # Ensure a valid, unique sheet name (Excel max 31 chars)
        tab_name = (tab_name or name)[:31]
        worksheet = workbook.create_sheet(tab_name)

        # Add column descriptions/units
        cols = fetch_columns(fp)
        colmeta = tab_meta.get("columns", {})
        for c_idx, column in enumerate(cols, 1):
            desc = colmeta.get(column, {}).get("description", column)
            unit = colmeta.get(column, {}).get("unit", "")
            worksheet.cell(row=1, column=c_idx, value=desc)
            worksheet.cell(row=2, column=c_idx, value=unit)

        # Append data rows
        for r in dataframe_to_rows(df, index=False):
            # dataframe_to_rows yields header first; we already wrote descriptions/units,
            # so skip the header row it produces.
            if r and r[0] == df.columns[0]:
                continue
            worksheet.append(r)

    # Save workbook
    fp_out = Path(
        DIR_LOG_REPORT / f"change_report {datetime.now().strftime('%Y-%m-%d')}.xlsx"
    )
    workbook.save(fp_out)
    empty_log_files()


# --- fetch_columns / fetch_tab_name (safe) ------------------------------------


def fetch_columns(variable):
    """
    Read reporting.yaml and return the columns for the variable.
    `variable` may be a Path or a string stem.
    """
    stem = variable.stem if hasattr(variable, "stem") else Path(variable).stem
    with open(LOG_REPORTING_FILEPATH, "r", encoding="utf-8") as stream:
        reporting = yaml.safe_load(stream)

    # Defensive: return listed columns or empty list
    cols = reporting.get(stem, {}).get("columns", {})
    return list(cols.keys())


def fetch_tab_name(variable):
    """
    Read reporting.yaml and return the tab name for the variable key (string).
    """
    key = variable if isinstance(variable, str) else str(variable)
    with open(LOG_REPORTING_FILEPATH, "r", encoding="utf-8") as stream:
        reporting = yaml.safe_load(stream)
    return reporting.get(key, {}).get("tab", key)


# --- convert_log_to_excel_file (hardened) -------------------------------------


def convert_log_to_excel_file(filepath):
    df = pd.read_csv(
        filepath, sep="|", header=None, on_bad_lines="skip", engine="python"
    )
    cols = fetch_columns(filepath)

    # Align column counts
    if df.shape[1] > len(cols):
        df = df.iloc[:, : len(cols)]
    elif df.shape[1] < len(cols):
        # pad with empty columns instead of pd.NA
        for _ in range(len(cols) - df.shape[1]):
            df[df.shape[1]] = None

    df.columns = cols

    # Make sure there is no pandas.NA left anywhere
    df = df.astype("object").where(df.notna(), None)

    return df
