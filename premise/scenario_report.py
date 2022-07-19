"""
This module export a summary of scenario to an Excel file.
"""

from pathlib import Path

import openpyxl
import yaml
from openpyxl.chart import AreaChart, LineChart, Reference
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from . import DATA_DIR

IAM_ELEC_VARS = DATA_DIR / "electricity" / "electricity_tech_vars.yml"
IAM_FUELS_VARS = DATA_DIR / "fuels" / "fuel_tech_vars.yml"
IAM_BIOMASS_VARS = DATA_DIR / "electricity" / "biomass_vars.yml"
IAM_CEMENT_VARS = DATA_DIR / "cement" / "cement_tech_vars.yml"
IAM_STEEL_VARS = DATA_DIR / "steel" / "steel_tech_vars.yml"
IAM_OTHER_VARS = DATA_DIR / "utils" / "report" / "other_vars.yaml"
GAINS_TO_IAM_FILEPATH = DATA_DIR / "GAINS_emission_factors" / "GAINStoREMINDtechmap.csv"
GNR_DATA = DATA_DIR / "cement" / "additional_data_GNR.csv"
IAM_CARBON_CAPTURE_VARS = DATA_DIR / "utils" / "carbon_capture_vars.yml"
REPORT_METADATA_FILEPATH = DATA_DIR / "utils" / "report" / "report.yaml"
VEHICLES_MAP = DATA_DIR / "transport" / "vehicles_map.yaml"
SECTORS = {
    "Population": (IAM_OTHER_VARS, ["Population"]),
    "GDP": (IAM_OTHER_VARS, ["GDP|PPP"]),
    "CO2": (IAM_OTHER_VARS, ["Emi|CO2"]),
    "Electricity - generation": IAM_ELEC_VARS,
    "Electricity (biom) - generation": IAM_BIOMASS_VARS,
    "Electricity - efficiency": IAM_ELEC_VARS,
    "Fuel - generation": IAM_FUELS_VARS,
    "Fuel - efficiency": IAM_FUELS_VARS,
    "Cement - generation": IAM_CEMENT_VARS,
    "Cement - efficiency": IAM_CEMENT_VARS,
    "Cement - CCS": (IAM_CARBON_CAPTURE_VARS, ["cement"]),
    "Steel - generation": IAM_STEEL_VARS,
    "Steel - efficiency": IAM_STEEL_VARS,
    "Steel - CCS": (IAM_CARBON_CAPTURE_VARS, ["steel"]),
    "Transport (cars)": (
        VEHICLES_MAP,
        ["BEV", "FCEV", "ICEV-d", "ICEV-g", "ICEV-p", "PHEV-d", "PHEV-p"],
    ),
    "Transport (buses)": (
        VEHICLES_MAP,
        ["BEV", "FCEV", "ICEV-d", "ICEV-g", "ICEV-p", "PHEV-d", "PHEV-p"],
    ),
    "Transport (trucks)": (
        VEHICLES_MAP,
        ["BEV", "FCEV", "ICEV-d", "ICEV-g", "ICEV-p", "PHEV-d", "PHEV-p"],
    ),
}


def get_variables(
    filepath,
):
    """
    Get the variables from a yaml file.
    :param filepath: path to the yaml file
    """
    with open(filepath, "r", encoding="utf-8") as stream:
        out = yaml.safe_load(stream)

    return list(out.keys())


def generate_summary_report(scenarios: list, filename: Path) -> None:
    """
    Generate a summary report of the scenarios.
    """

    with open(REPORT_METADATA_FILEPATH, "r", encoding="utf-8") as stream:
        metadata = yaml.safe_load(stream)

    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    for sector, filepath in SECTORS.items():

        if isinstance(filepath, tuple):
            filepath, variables = filepath
        else:
            variables = get_variables(filepath)

        worksheet = workbook.create_sheet(sector)

        col, row = (1, 1)

        worksheet.cell(
            column=col,
            row=row,
            value=metadata[sector]["expl_text"],
        )

        scenario_list = []

        last_col_used = 0

        for scenario_idx, scenario in enumerate(scenarios):

            if (scenario["model"], scenario["pathway"]) not in scenario_list:

                if "generation" in sector:
                    iam_data = scenario["iam data"].production_volumes
                elif "efficiency" in sector:
                    iam_data = scenario["iam data"].efficiency
                elif "CCS" in sector:
                    iam_data = scenario["iam data"].carbon_capture_rate * 100
                elif "car" in sector:
                    if scenario["iam data"].trsp_cars is not None:
                        iam_data = scenario["iam data"].trsp_cars.sum(
                            dim=["size", "construction_year"]
                        )
                        iam_data = iam_data.rename({"powertrain": "variables"}).T
                    else:
                        continue
                elif "bus" in sector:
                    if scenario["iam data"].trsp_buses is not None:
                        iam_data = scenario["iam data"].trsp_buses.sum(
                            dim=["size", "construction_year"]
                        )
                        iam_data = iam_data.rename({"powertrain": "variables"}).T
                    else:
                        continue
                elif "truck" in sector:
                    if scenario["iam data"].trsp_trucks is not None:
                        iam_data = scenario["iam data"].trsp_trucks.sum(
                            dim=["size", "construction_year"]
                        )
                        iam_data = iam_data.rename({"powertrain": "variables"}).T
                    else:
                        continue
                else:
                    iam_data = scenario["iam data"].other_vars

                if scenario_idx > 0:
                    col = last_col_used + metadata[sector]["offset"]

                row = 3

                worksheet.cell(
                    column=col,
                    row=row,
                    value=f"{scenario['model'].upper()} - {scenario['pathway'].upper()}",
                )
                worksheet.cell(column=col, row=row).font = Font(
                    bold=True, size=14, underline="single"
                )

                row += 2

                for region in scenario["iam data"].regions:
                    worksheet.cell(column=col, row=row, value=region)

                    row += 3

                    dataframe = iam_data.sel(
                        variables=[
                            v for v in variables if v in iam_data.variables.values
                        ],
                        region=region,
                        year=[y for y in iam_data.coords["year"].values if y <= 2100],
                    )

                    if len(dataframe) > 0:
                        dataframe = dataframe.to_dataframe("val")
                        dataframe = dataframe.unstack()["val"]
                        dataframe = dataframe.T
                        dataframe = dataframe.rename_axis(index=None)

                        data = dataframe_to_rows(dataframe)

                        counter = 0
                        for _, data_row in enumerate(data, 1):
                            if data_row != [None]:
                                for c_idx, value in enumerate(data_row, 1):
                                    worksheet.cell(
                                        row=row + counter,
                                        column=col + c_idx,
                                        value=value,
                                    )
                                    last_col_used = col + c_idx
                                counter += 1

                        values = Reference(
                            worksheet,
                            min_col=col + 2,
                            min_row=row,
                            max_col=col + c_idx,
                            max_row=row + counter - 1,
                        )
                        cats = Reference(
                            worksheet,
                            min_col=col + 1,
                            min_row=row + 1,
                            max_row=row + counter - 1,
                        )

                        if "generation" in sector:
                            chart = AreaChart(grouping="stacked")
                        elif "efficiency" in sector:
                            chart = LineChart()
                        elif "CCS" in sector:
                            chart = AreaChart()
                        elif "Transport" in sector:
                            chart = AreaChart(grouping="stacked")
                        else:
                            chart = LineChart()

                        chart.add_data(values, titles_from_data=True)
                        chart.set_categories(cats)
                        chart.title = f"{region} - {sector}"
                        chart.y_axis.title = metadata[sector]["label"]
                        chart.height = 8
                        chart.width = 16
                        chart.anchor = f"{get_column_letter(col + 2)}{row + 1}"
                        worksheet.add_chart(chart)

                        row += counter + 2

                scenario_list.append((scenario["model"], scenario["pathway"]))

    workbook.save(filename)
