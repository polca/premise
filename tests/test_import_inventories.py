# content of test_activity_maps.py
from pathlib import Path
from types import SimpleNamespace

import pytest
from openpyxl import load_workbook

from premise.filesystem_constants import INVENTORY_DIR
from premise.inventory_imports import (
    BaseInventoryImport,
    DefaultInventory,
    apply_migration_step,
    get_classification_entry,
    get_classifications,
)

FILEPATH_CARMA_INVENTORIES = INVENTORY_DIR / "lci-Carma-CCS.xlsx"
FILEPATH_CHP_INVENTORIES = INVENTORY_DIR / "lci-combined-heat-power-plant-CCS.xlsx"
FILEPATH_BIOFUEL_INVENTORIES = INVENTORY_DIR / "lci-biofuels.xlsx"
FILEPATH_BIOGAS_INVENTORIES = INVENTORY_DIR / "lci-biogas.xlsx"
FILEPATH_HYDROGEN_INVENTORIES = INVENTORY_DIR / "lci-hydrogen.xlsx"
FILEPATH_SYNFUEL_INVENTORIES = INVENTORY_DIR / "lci-synfuel.xlsx"
FILEPATH_SYNGAS_INVENTORIES = INVENTORY_DIR / "lci-syngas.xlsx"
FILEPATH_HYDROGEN_COAL_GASIFICATION_INVENTORIES = (
    INVENTORY_DIR / "lci-hydrogen-coal-gasification.xlsx"
)


def get_db():
    db = [
        {
            "code": "argsthyfujgyftdgr",
            "name": "fake activity",
            "reference product": "fake product",
            "location": "IAI Area, Africa",
            "unit": "kilogram",
            "exchanges": [
                {
                    "name": "fake activity",
                    "product": "fake product",
                    "amount": 1,
                    "type": "production",
                    "unit": "kilogram",
                    "input": ("dummy_db", "6543541"),
                },
                {
                    "name": "1,4-Butanediol",
                    "categories": ("air", "urban air close to ground"),
                    "amount": 1,
                    "type": "biosphere",
                    "unit": "kilogram",
                    "input": ("dummy_bio", "123"),
                },
            ],
        }
    ]
    version = "3.5"
    return db, version


class DummyInventoryImport(BaseInventoryImport):
    def load_inventory(self):
        return SimpleNamespace(data=[])


def get_replacement_source_db():
    return [
        {
            "code": "source",
            "name": "source process",
            "reference product": "source product",
            "location": "RoW",
            "unit": "kilogram",
            "exchanges": [
                {
                    "name": "source process",
                    "product": "source product",
                    "amount": 1,
                    "type": "production",
                    "unit": "kilogram",
                    "location": "RoW",
                },
                {
                    "name": "market for heat",
                    "product": "heat",
                    "amount": 3,
                    "type": "technosphere",
                    "unit": "megajoule",
                    "location": "RoW",
                },
                {
                    "name": "market for heat",
                    "product": "heat",
                    "amount": 0.5,
                    "type": "technosphere",
                    "unit": "megajoule",
                    "location": "Europe without Switzerland",
                },
                {
                    "name": "market group for electricity",
                    "product": "electricity, high voltage",
                    "amount": 2.5,
                    "type": "technosphere",
                    "unit": "kilowatt hour",
                    "location": "RER",
                },
                {
                    "name": "market group for electricity",
                    "product": "electricity, high voltage",
                    "amount": 9,
                    "type": "technosphere",
                    "unit": "kilowatt hour",
                    "location": "RoW",
                },
            ],
        }
    ]


def get_base_inventory_import(tmp_path, database):
    testpath = tmp_path / "testfile"
    testpath.write_text("")
    return DummyInventoryImport(
        database,
        version_in="3.8",
        version_out="3.8",
        path=testpath,
        system_model="cutoff",
    )


def test_file_exists():
    db, version = get_db()
    with pytest.raises(FileNotFoundError) as wrapped_error:
        BaseInventoryImport(
            db,
            version_in=version,
            version_out="3.8",
            path="testfile",
            system_model="cutoff",
        )
    assert wrapped_error.type == FileNotFoundError


def test_biosphere_dict():
    db, version = get_db()
    testpath = Path("testfile")
    open(testpath, "w")
    dbc = BaseInventoryImport(
        db, version_in=version, version_out="3.8", path=testpath, system_model="cutoff"
    )
    assert (
        dbc.biosphere_dict[
            ("1,4-Butanediol", "air", "urban air close to ground", "kilogram")
        ]
        == "38a622c6-f086-4763-a952-7c6b3b1c42ba"
    )

    testpath.unlink()


def test_biosphere_dict_2():
    db, version = get_db()
    testpath = Path("testfile")
    open(testpath, "w")
    dbc = BaseInventoryImport(
        db, version_in=version, version_out="3.8", path=testpath, system_model="cutoff"
    )

    for act in dbc.database:
        for exc in act["exchanges"]:
            if exc["type"] == "biosphere":
                assert (
                    dbc.biosphere_dict[
                        (
                            exc["name"],
                            exc["categories"][0],
                            exc["categories"][1],
                            exc["unit"],
                        )
                    ]
                    == "38a622c6-f086-4763-a952-7c6b3b1c42ba"
                )

    testpath.unlink()


def test_fill_data_gaps_matches_technosphere_exchange_location(tmp_path):
    dbc = get_base_inventory_import(tmp_path, get_replacement_source_db())
    exchange = {
        "name": "market for heat",
        "product": "heat",
        "location": "RoW",
        "amount": 0,
        "type": "technosphere",
        "unit": "megajoule",
        "replacement name": "source process",
        "replacement product": "source product",
        "replacement location": "RoW",
    }

    dbc.fill_data_gaps(exchange)

    assert exchange["amount"] == 3
    assert "replacement name" not in exchange
    assert "replacement product" not in exchange
    assert "replacement location" not in exchange


def test_fill_data_gaps_sums_technosphere_exchanges_without_location(tmp_path):
    dbc = get_base_inventory_import(tmp_path, get_replacement_source_db())
    exchange = {
        "name": "market for heat",
        "product": "heat",
        "amount": 0,
        "type": "technosphere",
        "unit": "megajoule",
        "replacement name": "source process",
        "replacement product": "source product",
        "replacement location": "RoW",
    }

    dbc.fill_data_gaps(exchange)

    assert exchange["amount"] == 3.5


def test_fill_data_gaps_market_group_fallback_matches_location(tmp_path):
    dbc = get_base_inventory_import(tmp_path, get_replacement_source_db())
    exchange = {
        "name": "market for electricity",
        "product": "electricity, high voltage",
        "location": "RER",
        "amount": 0,
        "type": "technosphere",
        "unit": "kilowatt hour",
        "replacement name": "source process",
        "replacement product": "source product",
        "replacement location": "RoW",
    }

    dbc.fill_data_gaps(exchange)

    assert exchange["amount"] == 2.5


def test_fill_data_gaps_applies_replacement_amount_multiplier(tmp_path):
    dbc = get_base_inventory_import(tmp_path, get_replacement_source_db())
    exchange = {
        "name": "market for heat",
        "product": "heat",
        "location": "RoW",
        "amount": 0,
        "type": "technosphere",
        "unit": "megajoule",
        "replacement name": "source process",
        "replacement product": "source product",
        "replacement location": "RoW",
        "replacement amount multiplier": 1.2,
    }

    dbc.fill_data_gaps(exchange)

    assert exchange["amount"] == pytest.approx(3.6)
    assert "replacement name" not in exchange
    assert "replacement product" not in exchange
    assert "replacement location" not in exchange
    assert "replacement amount multiplier" not in exchange


def test_fill_data_gaps_can_use_replacement_biosphere_source(tmp_path):
    testpath = tmp_path / "dummy.xlsx"
    testpath.write_text("")

    reference_db = [
        {
            "name": "source process",
            "reference product": "source product",
            "location": "RoW",
            "unit": "kilogram",
            "exchanges": [
                {
                    "name": "Carbon dioxide, fossil",
                    "categories": ("air", "high population density"),
                    "amount": 3,
                    "type": "biosphere",
                    "unit": "kilogram",
                },
                {
                    "name": "Carbon dioxide, fossil",
                    "categories": ("air", "lower stratosphere + upper troposphere"),
                    "amount": 2,
                    "type": "biosphere",
                    "unit": "kilogram",
                },
            ],
        }
    ]
    importer = DummyInventoryImport(
        reference_db,
        version_in="3.8",
        version_out="3.8",
        path=testpath,
        system_model="cutoff",
    )
    exchange = {
        "name": "carbon dioxide, captured at source",
        "product": "carbon dioxide, captured at source",
        "amount": 0,
        "type": "technosphere",
        "unit": "kilogram",
        "replacement name": "source process",
        "replacement product": "source product",
        "replacement location": "RoW",
        "replacement exchange type": "biosphere",
        "replacement exchange name": "Carbon dioxide, fossil",
        "replacement exchange unit": "kilogram",
        "replacement amount multiplier": 0.9,
    }

    importer.fill_data_gaps(exchange)

    assert exchange["amount"] == pytest.approx(4.5)
    assert "replacement name" not in exchange
    assert "replacement exchange type" not in exchange
    assert "replacement amount multiplier" not in exchange


def test_load_carma():
    db, version = get_db()
    carma = DefaultInventory(
        db,
        version_in="3.5",
        version_out="3.8",
        path=FILEPATH_CARMA_INVENTORIES,
        system_model="cutoff",
        keep_uncertainty_data=False,
    )
    assert len(carma.import_db.data) >= 81


def test_chp_ccs_inventory_embeds_upstream_chp_exchanges():
    workbook = load_workbook(FILEPATH_CHP_INVENTORIES, data_only=False)
    sheet = workbook.active
    activity_rows = [
        row
        for row in range(1, sheet.max_row + 1)
        if sheet.cell(row, 1).value == "Activity"
    ]

    assert len(activity_rows) == 8

    for index, start in enumerate(activity_rows):
        end = (
            activity_rows[index + 1]
            if index + 1 < len(activity_rows)
            else sheet.max_row + 1
        )
        header_row = next(
            row
            for row in range(start, end)
            if sheet.cell(row, 1).value == "name"
            and sheet.cell(row - 1, 1).value == "Exchanges"
        )
        headers = {
            sheet.cell(header_row, column).value: column
            for column in range(1, sheet.max_column + 1)
            if sheet.cell(header_row, column).value is not None
        }
        exchanges = [
            {key: sheet.cell(row, column).value for key, column in headers.items()}
            for row in range(header_row + 1, end)
            if sheet.cell(row, headers["type"]).value
            in {"production", "technosphere", "biosphere"}
        ]
        source_exchanges = [
            exchange
            for exchange in exchanges
            if exchange["type"] in {"technosphere", "biosphere"}
        ]
        capture_exchanges = [
            exchange
            for exchange in source_exchanges
            if exchange["type"] == "technosphere"
            and exchange["name"].startswith("carbon dioxide, captured")
        ]

        assert len(exchanges) > 30
        assert len(capture_exchanges) == 1
        assert all(exchange["amount"] == 0 for exchange in source_exchanges)
        assert all(exchange["replacement name"] for exchange in source_exchanges)
        assert capture_exchanges[0]["replacement exchange type"] == "biosphere"
        assert capture_exchanges[0]["replacement exchange name"] in {
            "Carbon dioxide, fossil",
            "Carbon dioxide, non-fossil",
        }
        assert not [
            exchange
            for exchange in source_exchanges
            if exchange["type"] == "technosphere"
            and exchange["name"].startswith("heat and power co-generation")
        ]
        assert not [
            exchange
            for exchange in source_exchanges
            if exchange["type"] == "biosphere"
            and exchange["name"]
            in {"Carbon dioxide, fossil", "Carbon dioxide, non-fossil"}
            and exchange["amount"] < 0
        ]


def test_load_biofuel():
    db, version = get_db()
    bio = DefaultInventory(
        db,
        version_in="3.7",
        version_out="3.8",
        path=FILEPATH_BIOFUEL_INVENTORIES,
        system_model="cutoff",
        keep_uncertainty_data=False,
    )
    assert len(bio.import_db.data) >= 150


def test_get_classifications_repairs_mojibake(tmp_path, monkeypatch):
    filepath = tmp_path / "classifications.csv"
    name = (
        "market for sawlog and veneer log, paran√° pine, "
        "measured as solid wood under bark"
    )
    product = "sawlog and veneer log, paran√° pine, measured as solid wood under bark"
    filepath.write_text(
        "\n".join(
            [
                "name,product,ISIC rev.4 ecoinvent,CPC",
                f'"{name}","{product}",0220:Logging,"3110: Wood"',
            ]
        ),
        encoding="utf-8",
    )

    get_classifications.cache_clear()
    monkeypatch.setattr("premise.inventory_imports.FILEPATH_CLASSIFICATIONS", filepath)

    classifications = get_classifications()
    classification = get_classification_entry(
        classifications,
        (
            "market for sawlog and veneer log, parana\u0301 pine, "
            "measured as solid wood under bark"
        ),
        "sawlog and veneer log, paraná pine, measured as solid wood under bark",
    )

    assert classification == {
        "ISIC rev.4 ecoinvent": "0220:Logging",
        "CPC": "3110: Wood",
    }

    get_classifications.cache_clear()


def test_fill_data_gaps_prefers_matching_supplier_location(tmp_path):
    testpath = tmp_path / "dummy.xlsx"
    testpath.write_text("")

    reference_db = [
        {
            "name": "pig iron production",
            "reference product": "pig iron",
            "location": "RER",
            "exchanges": [
                {
                    "name": "market for coke",
                    "product": "coke",
                    "location": "GLO",
                    "amount": 9.72,
                    "type": "technosphere",
                    "unit": "megajoule",
                },
                {
                    "name": "market for coke",
                    "product": "coke",
                    "location": "RoW",
                    "amount": 9.68,
                    "type": "technosphere",
                    "unit": "megajoule",
                },
            ],
        }
    ]

    importer = DummyInventoryImport(
        reference_db,
        version_in="3.8",
        version_out="3.8",
        path=testpath,
        system_model="cutoff",
    )

    exchange = {
        "name": "market for coke",
        "product": "coke",
        "location": "GLO",
        "unit": "megajoule",
        "type": "technosphere",
        "replacement name": "pig iron production",
        "replacement product": "pig iron",
        "replacement location": "RER",
    }

    importer.fill_data_gaps(exchange)

    assert exchange["amount"] == pytest.approx(9.72)
    assert "replacement name" not in exchange
    assert "replacement product" not in exchange
    assert "replacement location" not in exchange


def test_toggle_market_name_only_changes_market_prefix():
    embedded_market_name = (
        "heat, from municipal waste incineration to generic market for heat district "
        "or industrial, other than natural gas"
    )

    assert (
        BaseInventoryImport._toggle_market_name("market for hard coal")
        == "market group for hard coal"
    )
    assert (
        BaseInventoryImport._toggle_market_name("market group for hard coal")
        == "market for hard coal"
    )
    assert BaseInventoryImport._toggle_market_name(embedded_market_name) is None


def test_fill_dataset_data_gaps_replaces_migrated_split_markets(tmp_path):
    testpath = tmp_path / "dummy.xlsx"
    testpath.write_text("")

    reference_db = [
        {
            "name": "pig iron production",
            "reference product": "pig iron",
            "location": "RER",
            "unit": "kilogram",
            "exchanges": [
                {
                    "name": "market for coke",
                    "product": "coke",
                    "location": "RoW",
                    "amount": 9.724,
                    "type": "technosphere",
                    "unit": "megajoule",
                },
                {
                    "name": "market group for hard coal",
                    "product": "hard coal",
                    "location": "RER",
                    "amount": 0.15,
                    "type": "technosphere",
                    "unit": "kilogram",
                },
            ],
        }
    ]

    importer = DummyInventoryImport(
        reference_db,
        version_in="3.8",
        version_out="3.8",
        path=testpath,
        system_model="cutoff",
    )

    dataset = {
        "name": "pig iron production, blast furnace, with carbon capture and storage",
        "reference product": "pig iron",
        "location": "GLO",
        "unit": "kilogram",
        "exchanges": [
            {
                "name": "pig iron production, blast furnace, with carbon capture and storage",
                "product": "pig iron",
                "amount": 1,
                "type": "production",
                "unit": "kilogram",
            },
            {
                "name": "market for coke",
                "product": "coke",
                "location": "CN",
                "unit": "megajoule",
                "amount": 0,
                "type": "technosphere",
                "replacement name": "pig iron production",
                "replacement product": "pig iron",
                "replacement location": "RER",
            },
            {
                "name": "market for coke",
                "product": "coke",
                "location": "RoW",
                "unit": "megajoule",
                "amount": 0,
                "type": "technosphere",
                "replacement name": "pig iron production",
                "replacement product": "pig iron",
                "replacement location": "RER",
            },
            {
                "name": "market for hard coal",
                "product": "hard coal",
                "location": "DE",
                "unit": "kilogram",
                "amount": 0,
                "type": "technosphere",
                "replacement name": "pig iron production",
                "replacement product": "pig iron",
                "replacement location": "RER",
            },
            {
                "name": "market for hard coal",
                "product": "hard coal",
                "location": "FR",
                "unit": "kilogram",
                "amount": 0,
                "type": "technosphere",
                "replacement name": "pig iron production",
                "replacement product": "pig iron",
                "replacement location": "RER",
            },
        ],
    }

    importer.fill_dataset_data_gaps(dataset)

    techno = [exc for exc in dataset["exchanges"] if exc["type"] == "technosphere"]

    assert techno == [
        {
            "name": "market for coke",
            "product": "coke",
            "location": "RoW",
            "unit": "megajoule",
            "amount": pytest.approx(9.724),
            "type": "technosphere",
        },
        {
            "name": "market group for hard coal",
            "product": "hard coal",
            "location": "RER",
            "unit": "kilogram",
            "amount": pytest.approx(0.15),
            "type": "technosphere",
        },
    ]


def test_forward_migration_applies_in_memory_without_bw2io_datastore():
    class DummyImporter:
        data = [
            {
                "name": "old dataset",
                "reference product": "old product",
                "location": "GLO",
                "unit": "kilogram",
                "exchanges": [
                    {
                        "name": "old exchange",
                        "reference product": "old exchange product",
                        "location": "RER",
                        "unit": "megajoule",
                        "type": "technosphere",
                        "input": ("test", "old"),
                    }
                ],
            }
        ]

        def migrate(self, migration_name):
            raise AssertionError("bw2io Migration datastore should not be used")

    available = {
        ("3.11", "3.12"): {
            "replace": [
                {
                    "source": {
                        "name": "old dataset",
                        "reference product": "old product",
                        "location": "GLO",
                        "unit": "kilogram",
                    },
                    "target": {
                        "name": "new dataset",
                        "reference product": "new product",
                        "location": "RER",
                        "unit": "ton kilometer",
                    },
                },
                {
                    "source": {
                        "name": "old exchange",
                        "reference product": "old exchange product",
                        "location": "RER",
                        "unit": "megajoule",
                    },
                    "target": {
                        "name": "new exchange",
                        "reference product": "new exchange product",
                        "location": "CH",
                        "unit": "kilowatt hour",
                    },
                },
            ],
            "disaggregate": [],
        }
    }

    importer = DummyImporter()

    apply_migration_step(importer, "3.11", "3.12", "forward", available)

    assert importer.data[0]["name"] == "new dataset"
    assert importer.data[0]["reference product"] == "new product"
    assert importer.data[0]["location"] == "RER"
    assert importer.data[0]["unit"] == "kilogram"

    exchange = importer.data[0]["exchanges"][0]
    assert exchange["name"] == "new exchange"
    assert exchange["reference product"] == "new exchange product"
    assert exchange["location"] == "CH"
    assert exchange["unit"] == "megajoule"
    assert "input" not in exchange


def test_correct_product_field_uses_indexed_reference_product(tmp_path):
    testpath = tmp_path / "dummy.xlsx"
    testpath.write_text("")

    reference_db = [
        {
            "name": "shared supplier",
            "reference product": "first product",
            "location": "GLO",
            "unit": "kilogram",
            "exchanges": [],
        },
        {
            "name": "shared supplier",
            "reference product": "second product",
            "location": "GLO",
            "unit": "kilogram",
            "exchanges": [],
        },
    ]

    importer = DummyInventoryImport(
        reference_db,
        version_in="3.8",
        version_out="3.8",
        path=testpath,
        system_model="cutoff",
    )

    assert (
        importer.correct_product_field(
            ("shared supplier", "GLO", "kilogram", "second product")
        )
        == "second product"
    )
    assert (
        importer.correct_product_field(("shared supplier", "GLO", "kilogram", None))
        == "first product"
    )

    BaseInventoryImport.correct_product_field.cache_clear()


def test_legacy_hydrogen_market_exchange_is_restored_for_old_versions(tmp_path):
    importer = get_base_inventory_import(tmp_path, [])
    importer.import_db.data = [
        {
            "name": "consumer",
            "reference product": "consumer product",
            "location": "GLO",
            "unit": "kilogram",
            "exchanges": [
                {
                    "name": "market for hydrogen, gaseous, low pressure",
                    "reference product": "hydrogen, gaseous, low pressure",
                    "product": "hydrogen, gaseous, low pressure",
                    "location": "RER",
                    "unit": "kilogram",
                    "amount": 1,
                    "type": "technosphere",
                }
            ],
        }
    ]

    importer.adapt_hydrogen_market_exchanges_for_legacy_versions()

    exchange = importer.import_db.data[0]["exchanges"][0]
    assert exchange["name"] == "market for hydrogen, gaseous"
    assert exchange["reference product"] == "hydrogen, gaseous"
    assert exchange["product"] == "hydrogen, gaseous"
    assert exchange["location"] == "GLO"
