from pathlib import Path

import openpyxl
import xarray as xr

from premise import report
from premise.scenario_summary import (
    filter_sector_summary,
    get_sector_catalog,
    summarize_sector,
)


class FakeIAMData:
    pass


def _build_fake_iam_data():
    iam_data = FakeIAMData()
    iam_data.electricity_mix = True
    iam_data.production_volumes = xr.DataArray(
        [
            [[1.0, 2.0, 99.0], [3.0, 4.0, 100.0]],
            [[5.0, 6.0, 101.0], [7.0, 8.0, 102.0]],
        ],
        coords={
            "region": ["World", "Europe"],
            "variables": ["Biogas CHP", "Biomass CHP"],
            "year": [2020, 2030, 2110],
        },
        dims=("region", "variables", "year"),
    )
    iam_data.battery_mobile_scenarios = xr.DataArray(
        [
            [[0.6, 0.5, 0.4], [0.4, 0.5, 0.6]],
            [[0.7, 0.8, 0.9], [0.3, 0.2, 0.1]],
        ],
        coords={
            "scenario": ["Baseline", "High uptake"],
            "chemistry": ["NMC111", "LFP"],
            "year": [2020, 2030, 2110],
        },
        dims=("scenario", "chemistry", "year"),
    )
    iam_data.other_vars = xr.DataArray(
        [[[100.0, 120.0, 140.0]], [[50.0, 60.0, 70.0]]],
        coords={
            "region": ["World", "Europe"],
            "variables": ["population"],
            "year": [2020, 2030, 2110],
        },
        dims=("region", "variables", "year"),
    )
    return iam_data


def _build_fake_scenario(model="remind", pathway="SSP2-Base"):
    return {
        "model": model,
        "pathway": pathway,
        "filepath": str(Path("/tmp") / f"{model}_{pathway}.csv"),
        "iam data": _build_fake_iam_data(),
    }


def test_get_sector_catalog_attaches_metadata():
    catalog = get_sector_catalog()

    electricity = next(
        entry for entry in catalog if entry["id"] == "Electricity - generation"
    )

    assert electricity["unit"] == "Exajoules (EJ)"
    assert "Generated volumes of electricity" in electricity["explanation"]
    assert "Biogas CHP" in electricity["variables"]


def test_summarize_sector_returns_region_groups_and_trimmed_years():
    scenario = _build_fake_scenario()

    summary = summarize_sector([scenario], "Electricity - generation")

    assert summary["sector"] == "Electricity - generation"
    assert len(summary["scenarios"]) == 1

    scenario_summary = summary["scenarios"][0]
    assert scenario_summary["scenario_id"] == "remind::SSP2-Base::remind_SSP2-Base"
    assert scenario_summary["regions"] == ["Europe", "World"]
    assert scenario_summary["years"] == [2020, 2030]

    world_group = next(
        group for group in scenario_summary["groups"] if group["name"] == "World"
    )
    assert world_group["variables"] == ["Biogas CHP", "Biomass CHP"]
    assert world_group["years"] == [2020, 2030]
    assert world_group["series"][0]["points"] == [
        {"year": 2020, "value": 1},
        {"year": 2030, "value": 2},
    ]


def test_filter_sector_summary_filters_region_variable_and_year():
    summary = summarize_sector([_build_fake_scenario()], "Electricity - generation")

    filtered = filter_sector_summary(
        summary,
        regions=["World"],
        variables=["Biogas CHP"],
        year_end=2020,
    )

    assert filtered["regions"] == ["World"]
    assert filtered["variables"] == ["Biogas CHP"]
    assert filtered["years"] == [2020]
    assert len(filtered["scenarios"]) == 1
    assert len(filtered["scenarios"][0]["groups"]) == 1
    assert filtered["scenarios"][0]["groups"][0]["series"][0]["points"] == [
        {"year": 2020, "value": 1}
    ]


def test_filter_sector_summary_filters_battery_subscenario_by_group_name():
    summary = summarize_sector([_build_fake_scenario()], "Battery (mobile)")

    filtered = filter_sector_summary(summary, group_names=["Baseline"])

    assert filtered["subscenarios"] == ["Baseline"]
    assert len(filtered["scenarios"]) == 1
    assert len(filtered["scenarios"][0]["groups"]) == 1
    assert filtered["scenarios"][0]["groups"][0]["name"] == "Baseline"


def test_summarize_battery_sector_groups_by_subscenario():
    summary = summarize_sector([_build_fake_scenario()], "Battery (mobile)")

    assert summary["group_by"] == "subscenario"
    assert len(summary["scenarios"]) == 1

    scenario_summary = summary["scenarios"][0]
    assert scenario_summary["subscenarios"] == ["Baseline", "High uptake"]
    assert scenario_summary["years"] == [2020, 2030]

    baseline_group = next(
        group
        for group in scenario_summary["groups"]
        if group["subscenario"] == "Baseline"
    )
    assert baseline_group["variables"] == ["LFP", "NMC111"]
    assert baseline_group["series"][0]["points"] == [
        {"year": 2020, "value": 0.6},
        {"year": 2030, "value": 0.5},
    ]


def test_summarize_sector_skips_scenarios_with_missing_iam_attribute_payload():
    scenario = _build_fake_scenario()
    scenario["iam data"].production_volumes = None

    summary = summarize_sector([scenario], "Electricity - generation")

    assert summary["scenarios"] == []
    assert summary["years"] == []
    assert summary["regions"] == []


def test_generate_summary_report_renders_from_shared_summary_engine(
    monkeypatch, tmp_path
):
    monkeypatch.setattr(
        report,
        "load_report_metadata",
        lambda: {
            "Electricity - generation": {
                "label": "Exajoules (EJ)",
                "expl_text": "Generated volumes of electricity.",
                "offset": 3,
            }
        },
    )
    monkeypatch.setattr(
        report,
        "get_sector_catalog",
        lambda metadata=None: [{"id": "Electricity - generation"}],
    )
    monkeypatch.setattr(
        report,
        "summarize_sector",
        lambda scenarios, sector, metadata=None: {
            "sector": "Electricity - generation",
            "label": "Exajoules (EJ)",
            "explanation": "Generated volumes of electricity.",
            "offset": 3,
            "scenarios": [
                {
                    "model": "remind",
                    "pathway": "SSP2-Base",
                    "groups": [
                        {
                            "name": "World",
                            "group_type": "region",
                            "series": [
                                {
                                    "variable": "Biogas CHP",
                                    "points": [
                                        {"year": 2020, "value": 1},
                                        {"year": 2030, "value": 2},
                                    ],
                                },
                                {
                                    "variable": "Biomass CHP",
                                    "points": [
                                        {"year": 2020, "value": 3},
                                        {"year": 2030, "value": 4},
                                    ],
                                },
                            ],
                        }
                    ],
                }
            ],
        },
    )

    output_path = tmp_path / "summary.xlsx"
    report.generate_summary_report([], output_path, with_charts=False)

    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook["Electricity - generation"]

    assert worksheet["A1"].value == "Generated volumes of electricity."
    assert worksheet["A3"].value == "REMIND - SSP2-BASE"
    assert worksheet["A5"].value == "World"
    assert worksheet["B8"].value == "Biogas CHP"
    assert worksheet["A9"].value == 2020
    assert worksheet["B9"].value == 1


def test_generate_summary_report_matches_electricity_summary_payload(
    monkeypatch, tmp_path
):
    metadata = {
        "Electricity - generation": {
            "label": "Exajoules (EJ)",
            "expl_text": "Generated volumes of electricity.",
            "offset": 3,
        }
    }
    scenario = _build_fake_scenario()

    monkeypatch.setattr(report, "load_report_metadata", lambda: metadata)
    monkeypatch.setattr(
        report,
        "get_sector_catalog",
        lambda metadata=None: [{"id": "Electricity - generation"}],
    )

    summary = summarize_sector([scenario], "Electricity - generation", metadata=metadata)
    output_path = tmp_path / "summary-electricity.xlsx"
    report.generate_summary_report([scenario], output_path, with_charts=False)

    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook["Electricity - generation"]
    scenario_summary = summary["scenarios"][0]
    group = scenario_summary["groups"][0]

    assert worksheet["A1"].value == summary["explanation"]
    assert worksheet["A3"].value == "REMIND - SSP2-BASE"
    assert worksheet["A5"].value == group["name"]
    assert worksheet["B8"].value == group["series"][0]["variable"]
    assert worksheet["C8"].value == group["series"][1]["variable"]
    assert worksheet["A9"].value == group["series"][0]["points"][0]["year"]
    assert worksheet["B9"].value == group["series"][0]["points"][0]["value"]
    assert worksheet["C9"].value == group["series"][1]["points"][0]["value"]


def test_generate_summary_report_matches_battery_summary_payload(
    monkeypatch, tmp_path
):
    metadata = {
        "Battery (mobile)": {
            "label": "Share",
            "expl_text": "Mobile battery chemistry shares.",
            "offset": 2,
        }
    }
    scenario = _build_fake_scenario()

    monkeypatch.setattr(report, "load_report_metadata", lambda: metadata)
    monkeypatch.setattr(
        report,
        "get_sector_catalog",
        lambda metadata=None: [{"id": "Battery (mobile)"}],
    )

    summary = summarize_sector([scenario], "Battery (mobile)", metadata=metadata)
    output_path = tmp_path / "summary-battery.xlsx"
    report.generate_summary_report([scenario], output_path, with_charts=False)

    workbook = openpyxl.load_workbook(output_path)
    worksheet = workbook["Battery (mobile)"]
    scenario_summary = summary["scenarios"][0]
    group = scenario_summary["groups"][0]

    assert worksheet["A1"].value == summary["explanation"]
    assert worksheet["A3"].value == "REMIND - SSP2-BASE"
    assert worksheet["A5"].value == group["name"]
    assert worksheet["B8"].value == group["series"][0]["variable"]
    assert worksheet["A9"].value == group["series"][0]["points"][0]["year"]
    assert worksheet["B9"].value == group["series"][0]["points"][0]["value"]
