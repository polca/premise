from types import SimpleNamespace

import numpy as np
from openpyxl import load_workbook
import pytest
import xarray as xr

from premise.carbon_dioxide_removal import CarbonDioxideRemoval
from premise.data_collection import IAMDataCollection
from premise.filesystem_constants import INVENTORY_DIR

CDR_INVENTORY = INVENTORY_DIR / "lci-carbon-capture.xlsx"


def get_cdr_allocation_transform(
    database=None,
    cdr_volume=-25.0,
    co2_volume=75.0,
    kyoto_gases_volume=None,
    regions=("EUR",),
    ecoinvent_to_iam_loc=None,
):
    regions = list(regions)
    if np.isscalar(cdr_volume):
        cdr_values = np.full((1, len(regions), 1), cdr_volume)
    else:
        cdr_values = np.asarray(cdr_volume, dtype=float).reshape(1, len(regions), 1)
    if np.isscalar(co2_volume):
        co2_values = np.full((1, len(regions), 1), co2_volume)
    else:
        co2_values = np.asarray(co2_volume, dtype=float).reshape(1, len(regions), 1)
    other_values = [co2_values]
    other_variables = ["CO2"]
    if kyoto_gases_volume is not None:
        if np.isscalar(kyoto_gases_volume):
            kyoto_gases_values = np.full((1, len(regions), 1), kyoto_gases_volume)
        else:
            kyoto_gases_values = np.asarray(kyoto_gases_volume, dtype=float).reshape(
                1, len(regions), 1
            )
        other_values.append(kyoto_gases_values)
        other_variables.append("Kyoto Gases")

    cdr = object.__new__(CarbonDioxideRemoval)
    cdr.database = database or []
    cdr.model = "image"
    cdr.scenario = "SSP2"
    cdr.year = 2030
    cdr.regions = regions
    cdr.ecoinvent_to_iam_loc = ecoinvent_to_iam_loc or {"CH": "EUR"}
    cdr.cdr_map = {}
    cdr.iam_data = SimpleNamespace(
        cdr_technology_mix=True,
        cdr_energy_use=None,
        production_volumes=xr.DataArray(
            cdr_values,
            dims=("variables", "region", "year"),
            coords={
                "variables": ["enhanced rock weathering"],
                "region": regions,
                "year": [2030],
            },
        ),
        other_vars=xr.DataArray(
            np.concatenate(other_values, axis=0),
            dims=("variables", "region", "year"),
            coords={
                "variables": other_variables,
                "region": regions,
                "year": [2030],
            },
        ),
    )
    return cdr


def get_cdr_transform(
    electricity_efficiency=2.0,
    heat_efficiency=0.5,
    technology="direct air capture (solvent, gas heat) with storage",
):
    cdr = object.__new__(CarbonDioxideRemoval)
    cdr.year = 2030
    cdr.iam_data = SimpleNamespace(
        cdr_technology_efficiencies=xr.DataArray(
            np.array([[[[electricity_efficiency]], [[heat_efficiency]]]]),
            dims=("variables", "carrier", "region", "year"),
            coords={
                "variables": [technology],
                "carrier": ["electricity", "heat"],
                "region": ["EUR"],
                "year": [2030],
            },
        )
    )
    return cdr, technology


def test_cdr_allocation_share_uses_absolute_cdr_volume():
    cdr = get_cdr_allocation_transform(cdr_volume=-25.0, co2_volume=75.0)

    shares = cdr.calculate_cdr_allocation_shares()
    coverage_shares = cdr.calculate_cdr_allocation_coverage_shares()

    assert shares["EUR"] == pytest.approx(0.25)
    assert shares["World"] == pytest.approx(0.0)
    assert coverage_shares["EUR"]["co2"] == pytest.approx(0.25)
    assert coverage_shares["EUR"]["non_co2"] == pytest.approx(0.25)


def test_cdr_allocation_share_uses_remaining_cdr_for_non_co2_kyoto_gases():
    cdr = get_cdr_allocation_transform(
        cdr_volume=-100.0,
        co2_volume=-20.0,
        kyoto_gases_volume=30.0,
    )

    shares = cdr.calculate_cdr_allocation_coverage_shares()

    assert shares["EUR"]["co2"] == pytest.approx(1.0)
    assert shares["EUR"]["non_co2"] == pytest.approx(0.4)
    assert cdr.calculate_cdr_allocation_shares()["EUR"] == pytest.approx(1.0)


def test_cdr_allocation_share_uses_world_region_for_global_datasets():
    cdr_market = {
        "name": "market for carbon dioxide removal",
        "reference product": "carbon dioxide, captured and stored",
        "location": "World",
        "unit": "kilogram",
        "exchanges": [],
    }
    emitting_dataset = {
        "name": "global fossil emitting activity",
        "reference product": "product",
        "location": "GLO",
        "unit": "kilogram",
        "exchanges": [
            {
                "name": "Carbon dioxide, fossil",
                "amount": 12.0,
                "unit": "kilogram",
                "type": "biosphere",
            },
            {
                "name": "Carbon dioxide, non-fossil",
                "amount": 100.0,
                "unit": "kilogram",
                "type": "biosphere",
            },
        ],
    }
    cdr = get_cdr_allocation_transform(
        database=[cdr_market, emitting_dataset],
        cdr_volume=[25.0, 50.0],
        co2_volume=[75.0, 50.0],
        regions=("EUR", "World"),
        ecoinvent_to_iam_loc={"GLO": "World"},
    )

    cdr.allocate_cdr_to_greenhouse_gases()

    cdr_input = next(
        exc
        for exc in emitting_dataset["exchanges"]
        if exc["type"] == "technosphere"
        and exc["name"] == "market for carbon dioxide removal"
    )
    assert cdr_input["amount"] == pytest.approx(6.0)
    assert cdr_input["location"] == "World"
    assert emitting_dataset["log parameters"]["cdr allocation share"] == pytest.approx(
        0.5
    )
    assert emitting_dataset["log parameters"][
        "gross greenhouse gas emissions, kg CO2e"
    ] == pytest.approx(12.0)


def test_cdr_allocation_adds_regional_market_input_for_greenhouse_gases():
    cdr_market = {
        "name": "market for carbon dioxide removal",
        "reference product": "carbon dioxide, captured and stored",
        "location": "EUR",
        "unit": "kilogram",
        "regionalized": True,
        "exchanges": [
            {
                "name": "market for carbon dioxide removal",
                "product": "carbon dioxide, captured and stored",
                "location": "EUR",
                "amount": 1.0,
                "unit": "kilogram",
                "type": "production",
            }
        ],
    }
    emitting_dataset = {
        "name": "emitting activity",
        "reference product": "product",
        "location": "CH",
        "unit": "kilogram",
        "exchanges": [
            {
                "name": "emitting activity",
                "product": "product",
                "location": "CH",
                "amount": 1.0,
                "unit": "kilogram",
                "type": "production",
            },
            {
                "name": "Carbon dioxide, fossil",
                "amount": 10.0,
                "unit": "kilogram",
                "type": "biosphere",
            },
            {
                "name": "Methane, fossil",
                "amount": 1.0,
                "unit": "kilogram",
                "type": "biosphere",
            },
            {
                "name": "Tetrafluoromethane",
                "amount": 0.001,
                "unit": "kilogram",
                "type": "biosphere",
            },
            {
                "name": "Hexafluoroethane",
                "amount": 0.001,
                "unit": "kilogram",
                "type": "biosphere",
            },
            {
                "name": "1,1,1,2-Tetrafluoroethane",
                "amount": 0.001,
                "unit": "kilogram",
                "type": "biosphere",
            },
        ],
    }
    cdr = get_cdr_allocation_transform(
        database=[cdr_market, emitting_dataset],
        cdr_volume=-25.0,
        co2_volume=75.0,
    )

    cdr.allocate_cdr_to_greenhouse_gases()

    fossil_co2 = next(
        exc
        for exc in emitting_dataset["exchanges"]
        if exc["type"] == "biosphere" and exc["name"] == "Carbon dioxide, fossil"
    )
    cdr_input = next(
        exc
        for exc in emitting_dataset["exchanges"]
        if exc["type"] == "technosphere"
        and exc["name"] == "market for carbon dioxide removal"
    )
    methane = next(
        exc
        for exc in emitting_dataset["exchanges"]
        if exc["type"] == "biosphere" and exc["name"] == "Methane, fossil"
    )

    assert fossil_co2["amount"] == pytest.approx(10.0)
    assert methane["amount"] == pytest.approx(1.0)
    assert cdr_input["amount"] == pytest.approx(
        (10.0 + 29.8 + 0.001 * 7380.0 + 0.001 * 12400.0 + 0.001 * 1526.0) * 0.25
    )
    assert cdr_input["product"] == "carbon dioxide, captured and stored"
    assert cdr_input["location"] == "EUR"
    assert emitting_dataset["log parameters"][
        "gross greenhouse gas emissions, kg CO2e"
    ] == pytest.approx(61.106)
    assert emitting_dataset["log parameters"][
        "greenhouse gas emissions reduced by CDR, kg CO2e"
    ] == pytest.approx(61.106 * 0.25)
    assert emitting_dataset["log parameters"][
        "greenhouse gas emissions covered by CDR, kg CO2e"
    ] == pytest.approx(61.106 * 0.25)
    assert emitting_dataset["log parameters"][
        "remaining greenhouse gas emissions, kg CO2e"
    ] == pytest.approx(61.106 * 0.75)


def test_cdr_allocation_uses_separate_co2_and_non_co2_kyoto_gas_shares():
    cdr_market = {
        "name": "market for carbon dioxide removal",
        "reference product": "carbon dioxide, captured and stored",
        "location": "EUR",
        "unit": "kilogram",
        "regionalized": True,
        "exchanges": [],
    }
    emitting_dataset = {
        "name": "emitting activity with net-negative regional CO2",
        "reference product": "product",
        "location": "CH",
        "unit": "kilogram",
        "exchanges": [
            {
                "name": "Carbon dioxide, fossil",
                "amount": 10.0,
                "unit": "kilogram",
                "type": "biosphere",
            },
            {
                "name": "Methane, fossil",
                "amount": 1.0,
                "unit": "kilogram",
                "type": "biosphere",
            },
        ],
    }
    cdr = get_cdr_allocation_transform(
        database=[cdr_market, emitting_dataset],
        cdr_volume=-100.0,
        co2_volume=-20.0,
        kyoto_gases_volume=30.0,
    )

    cdr.allocate_cdr_to_greenhouse_gases()

    fossil_co2 = next(
        exc
        for exc in emitting_dataset["exchanges"]
        if exc["type"] == "biosphere" and exc["name"] == "Carbon dioxide, fossil"
    )
    methane = next(
        exc
        for exc in emitting_dataset["exchanges"]
        if exc["type"] == "biosphere" and exc["name"] == "Methane, fossil"
    )
    cdr_input = next(
        exc
        for exc in emitting_dataset["exchanges"]
        if exc["type"] == "technosphere"
        and exc["name"] == "market for carbon dioxide removal"
    )

    assert fossil_co2["amount"] == pytest.approx(10.0)
    assert methane["amount"] == pytest.approx(1.0)
    assert cdr_input["amount"] == pytest.approx(10.0 + 29.8 * 0.4)
    assert emitting_dataset["log parameters"]["new amount of fossil CO2"] == (
        pytest.approx(10.0)
    )
    assert emitting_dataset["log parameters"]["cdr allocation share, CO2"] == (
        pytest.approx(1.0)
    )
    assert emitting_dataset["log parameters"][
        "cdr allocation share, non-CO2 Kyoto gases"
    ] == pytest.approx(0.4)
    assert emitting_dataset["log parameters"][
        "CO2 emissions reduced by CDR, kg CO2e"
    ] == pytest.approx(10.0)
    assert emitting_dataset["log parameters"][
        "CO2 emissions covered by CDR, kg CO2e"
    ] == pytest.approx(10.0)
    assert emitting_dataset["log parameters"][
        "non-CO2 Kyoto gas emissions reduced by CDR, kg CO2e"
    ] == pytest.approx(29.8 * 0.4)
    assert emitting_dataset["log parameters"][
        "non-CO2 Kyoto gas emissions covered by CDR, kg CO2e"
    ] == pytest.approx(29.8 * 0.4)


def test_cdr_allocation_keeps_lognormal_fossil_co2_exchange_unchanged():
    cdr_market = {
        "name": "market for carbon dioxide removal",
        "reference product": "carbon dioxide, captured and stored",
        "location": "EUR",
        "unit": "kilogram",
        "regionalized": True,
        "exchanges": [],
    }
    emitting_dataset = {
        "name": "fully mitigated emitting activity",
        "reference product": "product",
        "location": "EUR",
        "unit": "kilogram",
        "exchanges": [
            {
                "name": "Carbon dioxide, fossil",
                "amount": 10.0,
                "unit": "kilogram",
                "type": "biosphere",
                "uncertainty type": 2,
                "loc": 2.302585092994046,
                "scale": 0.1,
            },
        ],
    }
    cdr = get_cdr_allocation_transform(
        database=[cdr_market, emitting_dataset],
        cdr_volume=100.0,
        co2_volume=0.0,
    )

    cdr.allocate_cdr_to_greenhouse_gases()

    fossil_co2 = next(
        exc
        for exc in emitting_dataset["exchanges"]
        if exc["type"] == "biosphere" and exc["name"] == "Carbon dioxide, fossil"
    )
    cdr_input = next(
        exc
        for exc in emitting_dataset["exchanges"]
        if exc["type"] == "technosphere"
        and exc["name"] == "market for carbon dioxide removal"
    )
    assert fossil_co2["amount"] == pytest.approx(10.0)
    assert fossil_co2["uncertainty type"] == 2
    assert fossil_co2["loc"] == pytest.approx(2.302585092994046)
    assert fossil_co2["scale"] == pytest.approx(0.1)
    assert cdr_input["amount"] == pytest.approx(10.0)


def test_afforestation_duplicate_iam_variable_is_split_by_region():
    cdr = object.__new__(CarbonDioxideRemoval)
    cdr.model = "image"
    cdr.cdr_map = {
        "afforestation, eucalyptus plantation": [],
        "afforestation, poplar plantation": [],
        "direct air capture (solvent, gas heat) with storage": [],
    }
    production_volumes = xr.DataArray(
        np.array(
            [
                [[10.0], [10.0], [20.0]],
                [[10.0], [10.0], [20.0]],
                [[5.0], [7.0], [12.0]],
            ]
        ),
        dims=("variables", "region", "year"),
        coords={
            "variables": [
                "afforestation, eucalyptus plantation",
                "afforestation, poplar plantation",
                "direct air capture (solvent, gas heat) with storage",
            ],
            "region": ["BRA", "CAN", "World"],
            "year": [2030],
        },
    )

    constrained = cdr._apply_cdr_regional_technology_constraints(production_volumes)

    assert constrained.sel(
        variables="afforestation, eucalyptus plantation", region="BRA"
    ).item() == pytest.approx(10.0)
    assert constrained.sel(
        variables="afforestation, eucalyptus plantation", region="CAN"
    ).item() == pytest.approx(0.0)
    assert constrained.sel(
        variables="afforestation, eucalyptus plantation", region="World"
    ).item() == pytest.approx(10.0)
    assert constrained.sel(
        variables="afforestation, poplar plantation", region="BRA"
    ).item() == pytest.approx(0.0)
    assert constrained.sel(
        variables="afforestation, poplar plantation", region="CAN"
    ).item() == pytest.approx(10.0)
    assert constrained.sel(
        variables="afforestation, poplar plantation", region="World"
    ).item() == pytest.approx(10.0)
    assert constrained.sel(
        variables="direct air capture (solvent, gas heat) with storage", region="BRA"
    ).item() == pytest.approx(5.0)


def test_cdr_duplicate_production_volume_is_split_by_energy_carrier_share():
    technologies = [
        "direct air capture (solvent, gas heat) with storage",
        "direct air capture (solvent, industrial steam heat) with storage",
        "direct air capture (solvent, heat pump) with storage",
        "direct air capture (solvent, hydrogen heat) with storage",
    ]
    carriers = ["gases", "heat", "electricity", "hydrogen"]

    cdr = object.__new__(CarbonDioxideRemoval)
    cdr.model = "remind"
    cdr.iam_data = SimpleNamespace(
        cdr_energy_use=xr.DataArray(
            np.array([[[30.0]], [[10.0]], [[40.0]], [[20.0]]]),
            dims=("variables", "region", "year"),
            coords={
                "variables": [
                    f"{technology} - {carrier}"
                    for technology, carrier in zip(technologies, carriers)
                ],
                "region": ["EUR"],
                "year": [2050],
            },
        )
    )
    production_volumes = xr.DataArray(
        np.full((4, 1, 1), 100.0),
        dims=("variables", "region", "year"),
        coords={"variables": technologies, "region": ["EUR"], "year": [2050]},
    )

    split = cdr._split_cdr_production_volumes_by_carrier(production_volumes)

    expected = {
        "direct air capture (solvent, gas heat) with storage": 30.0,
        "direct air capture (solvent, industrial steam heat) with storage": 10.0,
        "direct air capture (solvent, heat pump) with storage": 40.0,
        "direct air capture (solvent, hydrogen heat) with storage": 20.0,
    }
    for technology, amount in expected.items():
        assert split.sel(variables=technology, region="EUR", year=2050).item() == (
            pytest.approx(amount)
        )

    assert split.sel(region="EUR", year=2050).sum(dim="variables").item() == (
        pytest.approx(100.0)
    )


def test_regionalize_cdr_activities_keeps_mapped_activities_for_scaled_pass():
    primary = {
        "name": (
            "carbon dioxide, captured and stored, with a sorbent-based direct air "
            "capture system, 100ktCO2, with heat pump heat, and grid electricity"
        ),
        "reference product": "carbon dioxide, captured",
    }
    support = {
        "name": (
            "amine-based silica production, for sorbent-based direct air capture "
            "system"
        ),
        "reference product": "amine-based silica",
    }

    cdr = object.__new__(CarbonDioxideRemoval)
    cdr.model = "image"
    cdr.iam_data = SimpleNamespace(production_volumes=None)
    cdr.mapping = SimpleNamespace(
        generate_cdr_map=lambda model: {
            "direct air capture (sorbent, heat pump) with storage": [primary]
        },
        generate_sets_from_filters=lambda filters: {
            "direct air capture": [primary, support]
        },
    )
    cdr._apply_cdr_regional_technology_constraints = lambda production_volumes: (
        production_volumes
    )

    calls = []
    cdr.process_and_add_activities = lambda **kwargs: calls.append(kwargs)

    cdr.regionalize_cdr_activities()

    assert len(calls) == 2
    support_call, mapped_call = calls

    assert support_call["mapping"] == {
        "amine-based silica production, for sorbent-based direct air capture system": [
            support
        ]
    }
    assert mapped_call["mapping"] == cdr.cdr_map
    assert mapped_call["efficiency_adjustment_fn"] == cdr.adjust_cdr_efficiency


def test_support_filter_keeps_shared_unscaled_mapped_activities():
    dac = {
        "name": (
            "carbon dioxide, captured and stored, with a sorbent-based direct air "
            "capture system, 100ktCO2"
        ),
        "reference product": "carbon dioxide, captured",
    }
    shared_beccs = {
        "name": (
            "carbon dioxide, captured and stored, at wood burning power plant, "
            "pipeline 200km, storage 1000m"
        ),
        "reference product": "carbon dioxide, captured",
    }
    support = {
        "name": (
            "amine-based silica production, for sorbent-based direct air capture "
            "system"
        ),
        "reference product": "amine-based silica",
    }

    filtered = CarbonDioxideRemoval._exclude_mapped_cdr_activities_from_support(
        support_activities={"direct air capture": [dac, shared_beccs, support]},
        cdr_map={
            "direct air capture (sorbent, heat pump) with storage": [dac],
            "biomass heat generation, with CCS": [shared_beccs],
        },
        technologies={"direct air capture (sorbent, heat pump) with storage"},
    )

    assert filtered == {"direct air capture": [shared_beccs, support]}


def test_efficiency_aliases_mark_cdr_technologies_for_adjusted_pass():
    technologies = CarbonDioxideRemoval._get_efficiency_adjusted_technologies(
        cdr_mapping={
            "biomass power generation, with CCS": {
                "efficiency_use_aliases": {
                    "heat": {"image": "Efficiency|Electricity|Biomass|w/ CCS|1"}
                }
            },
            "biofuels, with CCS": {
                "energy_use_aliases": {
                    "heat": {
                        "image": (
                            "Secondary Energy|Consumption|Liquids|Biomass|"
                            "Ethanol|Woody|w/CCS"
                        )
                    }
                }
            },
            "biochar": {},
        },
        model="image",
    )

    assert technologies == {
        "biomass power generation, with CCS",
        "biofuels, with CCS",
    }


def test_cdr_efficiency_adjustment_scales_electricity_and_heat_separately():
    cdr, technology = get_cdr_transform()
    dataset = {
        "name": "carbon dioxide, captured and stored, with DAC",
        "reference product": "carbon dioxide, captured and stored",
        "location": "EUR",
        "unit": "kilogram",
        "exchanges": [
            {
                "name": "market group for electricity, medium voltage",
                "product": "electricity, medium voltage",
                "amount": 10.0,
                "type": "technosphere",
                "unit": "kilowatt hour",
            },
            {
                "name": "market for heat, district or industrial, natural gas",
                "product": "heat, district or industrial, natural gas",
                "amount": 8.0,
                "type": "technosphere",
                "unit": "megajoule",
            },
            {
                "name": "market for diesel, burned in agricultural machinery",
                "product": "diesel, burned in agricultural machinery",
                "amount": 2.0,
                "type": "technosphere",
                "unit": "megajoule",
            },
            {
                "name": "sorbent production",
                "product": "sorbent",
                "amount": 4.0,
                "type": "technosphere",
                "unit": "kilogram",
            },
            {
                "name": "Water",
                "amount": 7.0,
                "type": "biosphere",
                "unit": "cubic meter",
            },
            {
                "name": "Carbon dioxide, fossil",
                "amount": 3.0,
                "type": "biosphere",
                "unit": "kilogram",
            },
        ],
    }

    cdr.adjust_cdr_efficiency(dataset, technology)

    amounts = {exc["name"]: exc["amount"] for exc in dataset["exchanges"]}
    assert amounts["market group for electricity, medium voltage"] == pytest.approx(5.0)
    assert amounts[
        "market for heat, district or industrial, natural gas"
    ] == pytest.approx(12.0)
    assert amounts[
        "market for diesel, burned in agricultural machinery"
    ] == pytest.approx(3.0)
    assert amounts["sorbent production"] == pytest.approx(4.0)
    assert amounts["Water"] == pytest.approx(7.0)
    assert amounts["Carbon dioxide, fossil"] == pytest.approx(3.0)
    assert dataset["log parameters"][
        "electricity efficiency scaling factor"
    ] == pytest.approx(0.5)
    assert dataset["log parameters"]["heat efficiency scaling factor"] == pytest.approx(
        1.5
    )


def test_cdr_efficiencies_by_carrier_accept_energy_and_efficiency_indices():
    iam = object.__new__(IAMDataCollection)
    iam.year = 2050
    iam.min_year = 2020
    iam.filepath_iam_files = "test"
    data = xr.DataArray(
        np.array(
            [
                [
                    [0.4, 0.8],
                    [10.0, 20.0],
                    [100.0, 100.0],
                    [5.0, 5.0],
                ]
            ]
        ),
        dims=("region", "variables", "year"),
        coords={
            "region": ["WEU"],
            "variables": [
                "Efficiency|Electricity|Biomass|w/ CCS|1",
                "Carbon Removal|Geological Storage|Biomass|Liquids",
                "Secondary Energy|Consumption|Liquids|Biomass|Ethanol|Woody|w/CCS",
                "Carbon Removal|Geological Storage|Biomass|Electricity",
            ],
            "year": [2020, 2050],
        },
    )

    efficiencies = iam.get_iam_efficiencies_by_carrier(
        data=data,
        production_labels={
            "biomass power generation, with CCS": (
                "Carbon Removal|Geological Storage|Biomass|Electricity"
            ),
            "biofuels, with CCS": ("Carbon Removal|Geological Storage|Biomass|Liquids"),
        },
        energy_labels_by_carrier={
            "heat": {
                "biofuels, with CCS": (
                    "Secondary Energy|Consumption|Liquids|Biomass|Ethanol|Woody|w/CCS"
                )
            }
        },
        efficiency_labels_by_carrier={
            "heat": {
                "biomass power generation, with CCS": (
                    "Efficiency|Electricity|Biomass|w/ CCS|1"
                )
            }
        },
    )

    assert efficiencies.sel(
        variables="biomass power generation, with CCS",
        carrier="heat",
        region="WEU",
        year=2050,
    ).item() == pytest.approx(2.0)
    assert efficiencies.sel(
        variables="biofuels, with CCS",
        carrier="heat",
        region="WEU",
        year=2050,
    ).item() == pytest.approx(2.0)


def test_cdr_efficiency_index_aliases_accept_parser_lists():
    iam = object.__new__(IAMDataCollection)
    iam.year = 2050
    iam.min_year = 2020
    iam.filepath_iam_files = "test"
    data = xr.DataArray(
        np.array([[[0.4, 0.8]]]),
        dims=("region", "variables", "year"),
        coords={
            "region": ["WEU"],
            "variables": ["Efficiency|Electricity|Biomass|w/ CCS|1"],
            "year": [2020, 2050],
        },
    )

    efficiencies = iam.get_iam_efficiencies_by_carrier(
        data=data,
        production_labels={},
        energy_labels_by_carrier={},
        efficiency_labels_by_carrier={
            "heat": {
                "biomass power generation, with CCS": [
                    "Efficiency|Electricity|Biomass|w/ CCS|1"
                ]
            }
        },
    )

    assert efficiencies.sel(
        variables="biomass power generation, with CCS",
        carrier="heat",
        region="WEU",
        year=2050,
    ).item() == pytest.approx(2.0)


def test_dac_energy_lower_bounds_limit_scaled_solvent_heat_route():
    cdr, technology = get_cdr_transform(
        electricity_efficiency=10.0,
        heat_efficiency=10.0,
    )
    dataset = {
        "name": (
            "carbon dioxide, captured and stored, with a solvent-based direct air "
            "capture system, 1MtCO2"
        ),
        "reference product": "carbon dioxide, captured and stored",
        "location": "EUR",
        "unit": "kilogram",
        "exchanges": [
            {
                "name": "market group for electricity, medium voltage",
                "product": "electricity, medium voltage",
                "amount": 0.345,
                "type": "technosphere",
                "unit": "kilowatt hour",
            },
            {
                "name": "market for heat, district or industrial, natural gas",
                "product": "heat, district or industrial, natural gas",
                "amount": 9.18,
                "type": "technosphere",
                "unit": "megajoule",
            },
        ],
    }

    cdr.adjust_cdr_efficiency(dataset, technology)

    amounts = {exc["name"]: exc["amount"] for exc in dataset["exchanges"]}
    assert amounts["market group for electricity, medium voltage"] == pytest.approx(
        0.7 / 3.6
    )
    assert amounts[
        "market for heat, district or industrial, natural gas"
    ] == pytest.approx(5.3)
    assert dataset["log parameters"][
        "electricity lower-bound scaling factor"
    ] == pytest.approx(0.7 / (0.345 * 0.5 * 3.6))
    assert dataset["log parameters"][
        "heat lower-bound scaling factor"
    ] == pytest.approx(5.3 / (9.18 * 0.5))


def test_heat_pump_dac_lower_bound_converts_heat_floor_to_electricity():
    technology = "direct air capture (solvent, heat pump) with storage"
    cdr, _ = get_cdr_transform(
        electricity_efficiency=10.0,
        heat_efficiency=10.0,
        technology=technology,
    )
    dataset = {
        "name": (
            "carbon dioxide, captured and stored, with a solvent-based direct air "
            "capture system, 1MtCO2, with heat pump heat, and grid electricity"
        ),
        "reference product": "carbon dioxide, captured and stored",
        "location": "EUR",
        "unit": "kilogram",
        "exchanges": [
            {
                "name": "market group for electricity, medium voltage",
                "product": "electricity, medium voltage",
                "amount": 0.345,
                "type": "technosphere",
                "unit": "kilowatt hour",
            },
            {
                "name": "market group for electricity, medium voltage",
                "product": "electricity, medium voltage",
                "amount": 0.85,
                "type": "technosphere",
                "unit": "kilowatt hour",
            },
        ],
    }

    cdr.adjust_cdr_efficiency(dataset, technology)

    electricity_total = sum(exc["amount"] for exc in dataset["exchanges"])
    expected_lower_bound = 6.0 / 3.6
    assert electricity_total == pytest.approx(expected_lower_bound)
    assert dataset["log parameters"][
        "electricity lower bound (MJ/kg CO2)"
    ] == pytest.approx(0.7 + 5.3 / 3.0)
    assert dataset["log parameters"]["heat lower bound (MJ/kg CO2)"] is None
    assert dataset["log parameters"][
        "heat lower-bound scaling factor"
    ] == pytest.approx(1.0)
    assert dataset["log parameters"]["total lower bound (MJ/kg CO2)"] == pytest.approx(
        6.0
    )
    assert dataset["log parameters"][
        "total lower-bound scaling factor"
    ] == pytest.approx(6.0 / (0.7 + 5.3 / 3.0))


def test_sorbent_heat_pump_dac_lower_bound_keeps_total_energy_above_floor():
    technology = "direct air capture (sorbent, heat pump) with storage"
    cdr, _ = get_cdr_transform(
        electricity_efficiency=10.0,
        heat_efficiency=10.0,
        technology=technology,
    )
    dataset = {
        "name": (
            "carbon dioxide, captured and stored, with a sorbent-based direct air "
            "capture system, 100ktCO2, with heat pump heat, and grid electricity"
        ),
        "reference product": "carbon dioxide, captured and stored",
        "location": "EUR",
        "unit": "kilogram",
        "exchanges": [
            {
                "name": "market group for electricity, medium voltage",
                "product": "electricity, medium voltage",
                "amount": 0.7,
                "type": "technosphere",
                "unit": "kilowatt hour",
            },
            {
                "name": "market group for electricity, medium voltage",
                "product": "electricity, medium voltage",
                "amount": 11.9 / 3.6 / 3,
                "type": "technosphere",
                "unit": "kilowatt hour",
            },
            {
                "name": "Carbon dioxide, in air",
                "amount": 1.0,
                "type": "biosphere",
                "unit": "kilogram",
            },
        ],
    }

    cdr.adjust_cdr_efficiency(dataset, technology)

    electricity_total = sum(
        exc["amount"] for exc in dataset["exchanges"] if exc["type"] == "technosphere"
    )
    uptake = next(
        exc for exc in dataset["exchanges"] if exc["name"] == "Carbon dioxide, in air"
    )
    assert electricity_total == pytest.approx(4.0 / 3.6)
    assert uptake["amount"] == pytest.approx(1.0)
    assert dataset["log parameters"]["total lower bound (MJ/kg CO2)"] == pytest.approx(
        4.0
    )


def test_hydrogen_dac_lower_bound_uses_delivered_heat():
    technology = "direct air capture (solvent, hydrogen heat) with storage"
    cdr, _ = get_cdr_transform(
        electricity_efficiency=10.0,
        heat_efficiency=10.0,
        technology=technology,
    )
    dataset = {
        "name": (
            "carbon dioxide, captured and stored, with a solvent-based direct air "
            "capture system, 1MtCO2, with hydrogen heat, and grid electricity"
        ),
        "reference product": "carbon dioxide, captured and stored",
        "location": "EUR",
        "unit": "kilogram",
        "exchanges": [
            {
                "name": "market group for electricity, medium voltage",
                "product": "electricity, medium voltage",
                "amount": 0.345,
                "type": "technosphere",
                "unit": "kilowatt hour",
            },
            {
                "name": "market for hydrogen, gaseous, low pressure",
                "product": "hydrogen, gaseous, low pressure",
                "amount": 9.18 / (120 * 0.9),
                "type": "technosphere",
                "unit": "kilogram",
            },
        ],
    }

    cdr.adjust_cdr_efficiency(dataset, technology)

    amounts = {exc["name"]: exc["amount"] for exc in dataset["exchanges"]}
    assert amounts["market for hydrogen, gaseous, low pressure"] == pytest.approx(
        5.3 / (120 * 0.9)
    )
    assert dataset["log parameters"][
        "heat lower-bound scaling factor"
    ] == pytest.approx(5.3 / (9.18 * 0.5))


def get_inventory_activity(name):
    workbook = load_workbook(CDR_INVENTORY, data_only=True, read_only=True)
    worksheet = workbook["DAC"]
    try:
        rows = list(worksheet.iter_rows(values_only=True))
        for index, row in enumerate(rows):
            if row and row[0] == "Activity" and row[1] == name:
                activity_rows = []
                for activity_row in rows[index + 1 :]:
                    if activity_row and activity_row[0] == "Activity":
                        break
                    activity_rows.append(activity_row)

                comment = next(
                    row[1] for row in activity_rows if row and row[0] == "comment"
                )
                header_index = next(
                    idx
                    for idx, row in enumerate(activity_rows)
                    if row and row[0] == "name"
                )
                header = activity_rows[header_index]
                exchanges = [
                    {
                        key: value
                        for key, value in zip(header, row)
                        if key is not None and value is not None
                    }
                    for row in activity_rows[header_index + 1 :]
                    if row and row[0]
                ]
                return comment, exchanges
    finally:
        workbook.close()

    raise AssertionError(f"Activity not found: {name}")


@pytest.mark.parametrize(
    ("activity_name", "hydrogen_amount"),
    [
        (
            "carbon dioxide, captured and stored, with a solvent-based direct air "
            "capture system, 1MtCO2, with hydrogen heat, and grid electricity",
            9.18 / (120 * 0.9),
        ),
        (
            "carbon dioxide, captured and stored, with a sorbent-based direct air "
            "capture system, 100ktCO2, with hydrogen heat, and grid electricity",
            11.9 / (120 * 0.9),
        ),
    ],
)
def test_dac_hydrogen_proxy_inventories_replace_natural_gas_heat(
    activity_name, hydrogen_amount
):
    comment, exchanges = get_inventory_activity(activity_name)

    assert "Hydrogen heat proxy" in comment
    assert not any(
        exc["name"] == "market for heat, district or industrial, natural gas"
        for exc in exchanges
    )

    hydrogen = next(
        exc
        for exc in exchanges
        if exc["name"] == "market for hydrogen, gaseous, low pressure"
    )
    assert hydrogen["amount"] == pytest.approx(hydrogen_amount)
    assert hydrogen["location"] == "RER"
    assert hydrogen["unit"] == "kilogram"
    assert hydrogen["reference product"] == "hydrogen, gaseous, low pressure"


@pytest.mark.parametrize(
    ("activity_name", "input_name", "amount"),
    [
        (
            "carbon dioxide, captured and stored, by re/afforestation, eucalyptus",
            "hardwood forestry, eucalyptus ssp., planted forest management",
            1 / 1.35,
        ),
        (
            "carbon dioxide, captured and stored, by re/afforestation, willow",
            "willow production, short rotation coppice",
            1 / 1.76,
        ),
    ],
)
def test_afforestation_inventories_have_importable_forestry_amounts(
    activity_name, input_name, amount
):
    _, exchanges = get_inventory_activity(activity_name)

    production = next(exc for exc in exchanges if exc["type"] == "production")
    assert production["name"] == activity_name
    assert production["amount"] == pytest.approx(1)

    forestry = next(exc for exc in exchanges if exc["name"] == input_name)
    assert forestry["type"] == "technosphere"
    assert forestry["amount"] == pytest.approx(amount)


def test_wood_ccs_inventory_uses_volkart_wood_energy_penalty():
    comment, exchanges = get_inventory_activity(
        "carbon dioxide, captured and stored, at wood burning power plant, "
        "pipeline 200km, storage 1000m"
    )

    assert "host power plant" in comment
    assert "Volkart et al. (2013)" in comment
    assert "3.565 MJ additional wood-derived host energy" in comment
    assert "charcoal input" in comment

    assert not any(exc["name"] == "market for charcoal" for exc in exchanges)
    assert not any(
        exc["name"] == "market group for electricity, low voltage" for exc in exchanges
    )

    heat = next(
        exc
        for exc in exchanges
        if exc["name"] == "heat and power co-generation, wood chips, 6667 kW"
    )
    assert heat["amount"] == pytest.approx(6.0 / (0.187 * 0.9 / 0.1))
    assert heat["unit"] == "megajoule"
    assert (
        heat["reference product"]
        == "heat, district or industrial, other than natural gas"
    )

    mea = next(exc for exc in exchanges if exc["name"] == "market for monoethanolamine")
    assert mea["amount"] == pytest.approx(2.84e-4)

    spent_solvent = next(
        exc
        for exc in exchanges
        if exc["name"]
        == "treatment of spent solvent mixture, hazardous waste incineration"
    )
    assert spent_solvent["amount"] == pytest.approx(-2.27e-4)

    activated_carbon = next(
        exc
        for exc in exchanges
        if exc["name"] == "market for activated carbon, granular"
    )
    assert activated_carbon["amount"] == pytest.approx(8.26e-5)


def test_hydrogen_ccs_inventory_uses_inferred_antonini_heat_and_electricity():
    comment, exchanges = get_inventory_activity(
        "carbon dioxide, captured and stored, from a hydrogen production plant "
        "using steam methane reforming of biomethane"
    )

    assert "host SMR plant" in comment
    assert "methyldiethanolamine" in comment
    assert "0.11598 kWh per kg CO2 stored" in comment
    assert "0.948 MJ per kg CO2" in comment
    assert "former CEMCAP-derived 4.0556 MJ/kg CO2" in comment
    assert "Final CO2 compression electricity is omitted" in comment

    assert not any(exc["name"] == "market for monoethanolamine" for exc in exchanges)
    assert not any(
        exc["name"] == "market group for electricity, low voltage"
        and "final compression" in exc.get("comment", "").lower()
        for exc in exchanges
    )

    mdea = next(
        exc for exc in exchanges if exc["name"] == "market for methyldiethanolamine"
    )
    assert mdea["amount"] == pytest.approx(3.4e-5)
    assert mdea["reference product"] == "methyldiethanolamine"

    electricity = next(
        exc
        for exc in exchanges
        if exc["name"] == "market group for electricity, low voltage"
    )
    assert electricity["amount"] == pytest.approx(
        (0.005477565380988522 - (-0.000466178173455711)) / 0.05124742496104852
    )
    assert electricity["unit"] == "kilowatt hour"
    assert electricity["reference product"] == "electricity, low voltage"

    heat = next(
        exc
        for exc in exchanges
        if exc["name"]
        == "heat production, biomethane, at boiler condensing modulating <100kW"
    )
    assert heat["amount"] == pytest.approx(
        (0.68 - electricity["amount"] * 3.6) / (1 - 282 / 390)
    )
    assert heat["amount"] != pytest.approx((3.76 - 0.11) / 0.9)
    assert heat["unit"] == "megajoule"
    assert heat["reference product"] == "heat, central or small-scale, biomethane"
    assert "Inferred low-pressure MDEA reboiler heat" in heat["comment"]


def test_cement_mea_biogenic_cdr_variant_stores_one_kg_non_fossil_co2():
    comment, exchanges = get_inventory_activity(
        "carbon dioxide, captured and stored, at cement production plant, "
        "from non-fossil carbon dioxide, using monoethanolamine"
    )

    assert "non-fossil CO2 only" in comment
    assert "fossil CO2 co-captured" in comment
    assert "intentionally ignored" in comment
    assert "Carbon dioxide, in air" in comment

    production = next(exc for exc in exchanges if exc["type"] == "production")
    assert production["name"] == (
        "carbon dioxide, captured and stored, at cement production plant, "
        "from non-fossil carbon dioxide, using monoethanolamine"
    )
    assert production["amount"] == pytest.approx(1)
    assert production["reference product"] == "carbon dioxide, captured"

    storage = next(
        exc
        for exc in exchanges
        if exc["name"] == "carbon dioxide compression, transport and storage"
    )
    assert storage["amount"] == pytest.approx(1)

    uptake = next(exc for exc in exchanges if exc["name"] == "Carbon dioxide, in air")
    assert uptake["amount"] == pytest.approx(1)
    assert uptake["type"] == "biosphere"
    assert uptake["categories"] == "natural resource::in air"

    heat = next(
        exc
        for exc in exchanges
        if exc["name"] == "market for heat, district or industrial, natural gas"
    )
    assert heat["amount"] == pytest.approx(4.055555555555555)

    electricity = [
        exc
        for exc in exchanges
        if exc["name"] == "market group for electricity, low voltage"
    ]
    assert sum(exc["amount"] for exc in electricity) == pytest.approx(
        0.03149802890932983 + 0.020607752956636 + 0.0963370565045992
    )
    assert not any(exc["name"] == "Carbon dioxide, fossil" for exc in exchanges)
    assert not any(exc["name"] == "Carbon dioxide, non-fossil" for exc in exchanges)


def test_fermentation_ccs_inventory_has_no_solvent_or_extra_capture_energy():
    comment, exchanges = get_inventory_activity(
        "carbon dioxide, captured and stored, from a biomass fermentation plant"
    )

    assert "already high-purity fermentation CO2 stream" in comment
    assert "no solvent, sorbent or regeneration heat" in comment
    assert "omitted here to avoid double-counting compression" in comment

    assert not any("electricity" in exc.get("name", "").lower() for exc in exchanges)
    assert not any("heat" in exc.get("name", "").lower() for exc in exchanges)
    assert not any("ethanolamine" in exc.get("name", "").lower() for exc in exchanges)


@pytest.mark.parametrize(
    "activity_name",
    [
        (
            "carbon dioxide, captured and stored, at wood burning power plant, "
            "pipeline 200km, storage 1000m"
        ),
        (
            "carbon dioxide, captured and stored, from a hydrogen production plant "
            "using steam methane reforming of biomethane"
        ),
        "carbon dioxide, captured and stored, from a biomass fermentation plant",
        (
            "carbon dioxide, captured and stored, at cement production plant, from "
            "non-fossil carbon dioxide, using monoethanolamine"
        ),
    ],
)
def test_image_scaled_cdr_inventories_have_no_direct_combustion_ghg_flows(
    activity_name,
):
    _, exchanges = get_inventory_activity(activity_name)
    direct_combustion_ghg = {
        "Carbon dioxide, fossil",
        "Carbon dioxide, non-fossil",
        "Methane, fossil",
        "Methane, non-fossil",
        "Methane",
        "Dinitrogen monoxide",
    }

    assert not any(
        exc["type"] == "biosphere" and exc["name"] in direct_combustion_ghg
        for exc in exchanges
    )
    assert any(
        exc["name"] == "carbon dioxide compression, transport and storage"
        for exc in exchanges
    )
