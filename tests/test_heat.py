from pathlib import Path
from types import SimpleNamespace

import pytest
import xarray as xr
from openpyxl import load_workbook

from premise.activity_maps import InventorySet
from premise.heat import (
    BUILDING_LEGACY_INPUTS,
    BUILDINGS_MARKET,
    Heat,
    _update_heat,
)
from premise.inventory_imports import DefaultInventory


def make_electric_supplier(amount, name="heat production, test"):
    return {
        "name": name,
        "reference product": "heat, district or industrial",
        "location": "CH",
        "unit": "megajoule",
        "exchanges": [
            {
                "name": "market for electricity, medium voltage",
                "product": "electricity, medium voltage",
                "location": "CH",
                "unit": "kilowatt hour",
                "amount": amount,
                "type": "technosphere",
            }
        ],
    }


def test_inventory_derived_electric_boiler_factor():
    heat = object.__new__(Heat)
    supplier = make_electric_supplier(0.28058361391694725)
    assert heat._supplier_conversion_factor(
        supplier, "electric_boiler"
    ) == pytest.approx(0.99)


def test_inventory_derived_heat_pump_factor():
    heat = object.__new__(Heat)
    supplier = make_electric_supplier(0.0805556)
    assert heat._supplier_conversion_factor(supplier, "heat_pump") == pytest.approx(
        3.4482746, rel=1e-6
    )


def test_electric_heat_factor_rejects_missing_positive_input():
    heat = object.__new__(Heat)
    with pytest.raises(ValueError, match="No positive electricity input"):
        heat._supplier_conversion_factor(make_electric_supplier(0), "electric_boiler")


def test_combustion_factor_recognizes_supplier_activity_name():
    heat = object.__new__(Heat)
    heat.database = [
        {
            "exchanges": [
                {
                    "name": "biomethane, gaseous, 5 bar, from sewage sludge fermentation, at fuelling station",
                    "product": "biomethane, high pressure",
                    "unit": "kilogram",
                    "amount": 0.02,
                    "type": "technosphere",
                }
            ]
        }
    ]
    heat.fuels_specs = {"methane, from biomass": {"lhv": {"value": 47.5}}}
    heat.fuel_map_reverse = heat._build_fuel_map_reverse(
        {
            "methane, from biomass": [
                {
                    "name": "biomethane production, high pressure, vehicle grade",
                    "reference product": "biomethane, high pressure",
                }
            ]
        }
    )
    supplier = {
        "name": "heat production, biomethane boiler",
        "reference product": "heat",
        "location": "CH",
        "unit": "megajoule",
        "exchanges": [
            {
                "name": "biomethane, gaseous, 5 bar, from sewage sludge fermentation, at fuelling station",
                "product": "biomethane, high pressure",
                "location": "CH",
                "unit": "kilogram",
                "amount": 0.02,
                "type": "technosphere",
            },
            {
                "name": "market for electricity, low voltage",
                "product": "electricity, low voltage",
                "location": "CH",
                "unit": "kilowatt hour",
                "amount": 0.001,
                "type": "technosphere",
            },
        ],
    }

    assert heat._supplier_conversion_factor(supplier, "combustion") == pytest.approx(
        1 / (0.02 * 47.5)
    )


def test_combustion_factor_recognizes_cleft_timber_as_wood_fuel():
    heat = object.__new__(Heat)
    heat.database = [
        {
            "exchanges": [
                {
                    "name": "market for cleft timber, green, measured as dry mass",
                    "product": "cleft timber, green, measured as dry mass",
                    "unit": "kilogram",
                    "amount": 0.1,
                    "type": "technosphere",
                }
            ]
        }
    ]
    heat.fuels_specs = {"wood chips": {"lhv": {"value": 16.2}}}
    heat.fuel_map_reverse = heat._build_fuel_map_reverse({})
    supplier = {
        "name": "heat production, mixed logs, at wood heater 6kW",
        "reference product": "heat",
        "location": "CH",
        "unit": "megajoule",
        "exchanges": heat.database[0]["exchanges"],
    }

    assert heat._supplier_conversion_factor(supplier, "combustion") == pytest.approx(
        1 / 1.62
    )


def test_combustion_factor_recognizes_fossil_methanol_energy_content():
    heat = object.__new__(Heat)
    exchange = {
        "name": "methanol production, from coal gasification",
        "product": "methanol",
        "unit": "kilogram",
        "amount": 0.05,
        "type": "technosphere",
    }
    heat.database = [{"exchanges": [exchange]}]
    heat.fuels_specs = {"methanol, from wood": {"lhv": {"value": 19.9}}}
    heat.fuel_map_reverse = heat._build_fuel_map_reverse({})
    supplier = {
        "name": "heat production, methanol boiler",
        "reference product": "heat",
        "location": "CH",
        "unit": "megajoule",
        "exchanges": [exchange],
    }

    assert heat._supplier_conversion_factor(supplier, "combustion") == pytest.approx(
        1 / (0.05 * 19.9)
    )


def test_delivered_heat_uses_supplier_weighted_regional_factor(monkeypatch):
    heat = object.__new__(Heat)
    heat.regions = ["WEU"]
    heat.year = 2030
    heat.diagnostics = {}
    heat.heat_metadata = {
        "technology": {
            "energy_basis": "final_energy",
            "conversion": "combustion",
        }
    }
    suppliers = [
        {"name": "one", "location": "A", "share": 0.25},
        {"name": "two", "location": "B", "share": 0.75},
    ]
    monkeypatch.setattr(
        heat, "_select_suppliers", lambda technology, region: (suppliers, "contained")
    )
    monkeypatch.setattr(
        heat,
        "_supplier_conversion_factor",
        lambda supplier, conversion: {"one": 0.5, "two": 1.0}[supplier["name"]],
    )
    raw = xr.DataArray(
        [[[2.0, 4.0]]],
        dims=("variables", "region", "year"),
        coords={"variables": ["technology"], "region": ["WEU"], "year": [2020, 2030]},
        attrs={"residual": {"technology": False}},
    )

    delivered = heat.convert_to_delivered_heat(raw, "industrial_end_use")
    assert delivered.values.tolist() == [[[1.75, 3.5]]]
    assert heat.diagnostics["industrial_end_use"]["conversion factors"][0][
        "factor"
    ] == pytest.approx(0.875)


def test_supplier_weighting_preserves_inventory_exchanges():
    heat = object.__new__(Heat)
    heat.iam_to_ecoinvent_loc = {"WEU": ["CH", "FR"]}
    heat.heat_techs = {
        "technology": [
            {
                "name": "supplier",
                "reference product": "heat",
                "location": "CH",
                "unit": "megajoule",
                "production volume": 1,
                "exchanges": [{"name": "fuel one"}],
            },
            {
                "name": "supplier",
                "reference product": "heat",
                "location": "FR",
                "unit": "megajoule",
                "production volume": 3,
                "exchanges": [{"name": "fuel two"}],
            },
        ]
    }

    suppliers, fallback = heat._select_suppliers("technology", "WEU")

    assert fallback == "contained ecoinvent location"
    assert [supplier["share"] for supplier in suppliers] == [0.25, 0.75]
    assert [supplier["exchanges"][0]["name"] for supplier in suppliers] == [
        "fuel one",
        "fuel two",
    ]


def test_model_without_any_heat_layer_is_left_unchanged():
    database = [{"name": "untouched"}]
    iam_data = SimpleNamespace(
        buildings_heat_end_use=None,
        industrial_heat_end_use=None,
        secondary_heat_supply=None,
        heat_diagnostics={"buildings_end_use": {"available": False}},
    )
    scenario = {
        "database": database,
        "iam data": iam_data,
        "model": "gcam",
        "pathway": "legacy",
        "year": 2050,
    }
    result = _update_heat(scenario, "3.10", "cutoff")
    assert result["database"] is database
    assert result["database"] == [{"name": "untouched"}]


def test_relink_excludes_new_dataset_codes(monkeypatch):
    legacy_exchange = {
        "name": "old heat",
        "product": "old product",
        "location": "CH",
        "unit": "megajoule",
        "amount": 1,
        "type": "technosphere",
        "input": ("db", "old"),
    }
    external = {
        "name": "consumer",
        "reference product": "service",
        "location": "CH",
        "unit": "unit",
        "code": "external",
        "exchanges": [dict(legacy_exchange)],
    }
    generated = {
        **external,
        "name": "generated consumer",
        "code": "generated",
        "exchanges": [dict(legacy_exchange)],
    }
    heat = object.__new__(Heat)
    heat.database = [external, generated]
    heat.created_dataset_codes = {"generated"}
    heat.regions = ["WEU"]
    heat.ecoinvent_to_iam_loc = {"CH": "WEU"}
    monkeypatch.setattr(heat, "is_in_index", lambda candidate, location: True)

    heat.relink_heat_markets(
        [{"name": "old heat", "reference product": "old product"}],
        {"name": "new heat", "reference product": "new product"},
    )

    rewritten = external["exchanges"][0]
    assert rewritten["name"] == "new heat"
    assert rewritten["product"] == "new product"
    assert rewritten["location"] == "WEU"
    assert "input" not in rewritten
    assert generated["exchanges"][0] == legacy_exchange


@pytest.mark.parametrize(
    "legacy_name,legacy_product",
    [
        (
            "market group for heat, central or small-scale, biomethane",
            "heat, central or small-scale, biomethane",
        ),
        (
            "market group for heat, central or small-scale, natural gas",
            "heat, central or small-scale, natural gas",
        ),
        (
            "market group for heat, central or small-scale, other than natural gas",
            "heat, central or small-scale, other than natural gas",
        ),
    ],
)
def test_building_legacy_market_groups_are_relinked(
    monkeypatch, legacy_name, legacy_product
):
    exchange = {
        "name": legacy_name,
        "product": legacy_product,
        "location": "RER",
        "unit": "megajoule",
        "amount": 0.5,
        "type": "technosphere",
        "input": ("db", "legacy"),
    }
    consumer = {
        "name": "building heat consumer",
        "reference product": "service",
        "location": "CH",
        "unit": "unit",
        "code": "consumer",
        "exchanges": [exchange],
    }
    heat = object.__new__(Heat)
    heat.database = [consumer]
    heat.created_dataset_codes = set()
    heat.regions = ["NEU"]
    heat.ecoinvent_to_iam_loc = {"CH": "NEU"}
    monkeypatch.setattr(heat, "is_in_index", lambda candidate, location: True)

    heat.relink_heat_markets(BUILDING_LEGACY_INPUTS, BUILDINGS_MARKET)

    rewritten = consumer["exchanges"][0]
    assert rewritten["name"] == BUILDINGS_MARKET["name"]
    assert rewritten["product"] == BUILDINGS_MARKET["reference product"]
    assert rewritten["location"] == "NEU"
    assert rewritten["amount"] == 0.5
    assert "input" not in rewritten


def test_generated_heat_cycle_is_rejected():
    first = {
        "name": "heat one",
        "reference product": "heat",
        "location": "WEU",
        "unit": "megajoule",
        "code": "one",
        "exchanges": [
            {
                "name": "heat two",
                "product": "heat",
                "location": "WEU",
                "unit": "megajoule",
                "amount": 1,
                "type": "technosphere",
            }
        ],
    }
    second = {
        **first,
        "name": "heat two",
        "code": "two",
        "exchanges": [
            {
                "name": "heat one",
                "product": "heat",
                "location": "WEU",
                "unit": "megajoule",
                "amount": 1,
                "type": "technosphere",
            }
        ],
    }
    heat = object.__new__(Heat)
    heat.database = [first, second]
    heat.created_dataset_codes = {"one", "two"}
    with pytest.raises(ValueError, match="Circular dependency"):
        heat.assert_no_heat_cycles()


@pytest.mark.parametrize(
    "filename,activity,exchange_name,amount",
    [
        (
            "lci-electric-boiler-industrial.xlsx",
            "heat production, electric boiler, industrial",
            "market for electricity, medium voltage",
            0.28058361391694725,
        ),
        (
            "lci-nuclear-heat.xlsx",
            "heat production, nuclear cogeneration",
            "electricity production, nuclear, Evolutionary Power Reactor (EPR)",
            0.05958528990799713,
        ),
    ],
)
def test_new_heat_inventory_workbooks(filename, activity, exchange_name, amount):
    path = (
        Path(__file__).parents[1]
        / "premise"
        / "data"
        / "additional_inventories"
        / filename
    )
    sheet = load_workbook(path, data_only=True).active
    assert sheet["B3"].value == activity
    rows = {
        sheet.cell(row, 1).value: sheet.cell(row, 2).value
        for row in range(1, sheet.max_row + 1)
    }
    assert rows[exchange_name] == pytest.approx(amount)


@pytest.mark.parametrize(
    "filename,activity",
    [
        (
            "lci-electric-boiler-industrial.xlsx",
            "heat production, electric boiler, industrial",
        ),
        ("lci-nuclear-heat.xlsx", "heat production, nuclear cogeneration"),
    ],
)
def test_new_heat_inventory_workbooks_are_importable(filename, activity):
    path = (
        Path(__file__).parents[1]
        / "premise"
        / "data"
        / "additional_inventories"
        / filename
    )
    inventory = DefaultInventory(
        database=[],
        version_in="3.10",
        version_out="3.10",
        path=path,
        system_model="cutoff",
        keep_uncertainty_data=False,
    )
    assert any(dataset["name"] == activity for dataset in inventory.import_db.data)


def test_new_heat_inventories_are_in_message_activity_mapping():
    inventory_dir = (
        Path(__file__).parents[1] / "premise" / "data" / "additional_inventories"
    )
    database = []
    for filename in (
        "lci-electric-boiler-industrial.xlsx",
        "lci-nuclear-heat.xlsx",
    ):
        imported = DefaultInventory(
            database=[],
            version_in="3.10",
            version_out="3.10",
            path=inventory_dir / filename,
            system_model="cutoff",
            keep_uncertainty_data=False,
        )
        database.extend(imported.import_db.data)

    mapping = InventorySet(database=database).generate_heat_map(model="message")
    assert mapping["heat, industrial, from electric boiler"]
    assert mapping["heat, secondary, from electricity"]
    assert mapping["heat, secondary, from nuclear"]
